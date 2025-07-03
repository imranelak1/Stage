import asyncio
import re
import uuid
from fastapi import FastAPI, HTTPException, status, WebSocket, WebSocketDisconnect, Cookie, Response, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, StreamingResponse
from pydantic import BaseModel
import secrets
from typing import Optional, List, Dict, Any
import hashlib
import json
from datetime import datetime, time, timezone, timedelta
from db_connection import Database
from pymysql.err import Error as PyMySQLError
import pandas as pd
from io import BytesIO
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import time as time_module
import logging
import traceback
from logger_config import logger


app = FastAPI(
    title="T3Shield API",
    description="API for T3Shield exam monitoring system",
    version="1.0.0"
)

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],  # Explicitly allow your frontend URLs
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["Content-Type", "X-Total-Count"],
)

# Add request logging middleware
@app.middleware("http")
async def log_requests(request: Request, call_next):
    start_time = time_module.time()
    response = await call_next(request)
    process_time = time_module.time() - start_time
    
    logger.info(f"{request.method} {request.url.path} - {response.status_code} - {process_time:.3f}s")
    return response

# Mount static files
app.mount("/static", StaticFiles(directory="front-end"), name="static")

# Database instance
db = Database()

# GMT+1 timezone (Morocco Standard Time)
GMT_PLUS_1 = timezone(timedelta(hours=1))

def get_morocco_time():
    """Get current time in GMT+1 (Morocco timezone)"""
    return datetime.now(GMT_PLUS_1)

def get_morocco_time_str():
    """Get current time in GMT+1 as string format YYYY-MM-DD HH:MM:SS"""
    return get_morocco_time().strftime("%Y-%m-%d %H:%M:%S")

def get_morocco_date_str():
    """Get current date in GMT+1 as string format YYYY-MM-DD"""
    return get_morocco_time().strftime("%Y-%m-%d")

def english_to_french_day(english_day):
    """Convert English day names to French"""
    day_mapping = {
        'Monday': 'Lundi',
        'Tuesday': 'Mardi', 
        'Wednesday': 'Mercredi',
        'Thursday': 'Jeudi',
        'Friday': 'Vendredi',
        'Saturday': 'Samedi',
        'Sunday': 'Dimanche'
    }
    return day_mapping.get(english_day, english_day)


class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: int
    full_name: str
    username: str
    profile_id: int
    profile_name: str

active_sessions = {}

# Dictionary to store active chef centre sessions
chef_centre_sessions = {}

def cleanup_expired_sessions():
    """Clean up expired chef centre sessions"""
    current_time = get_morocco_time()
    expired_sessions = []
    
    for token, session_data in chef_centre_sessions.items():
        login_time = session_data.get("login_time", current_time)
        # Sessions expire after 8 hours
        if (current_time - login_time).total_seconds() > 3600 * 8:
            expired_sessions.append(token)
    
    for token in expired_sessions:
        del chef_centre_sessions[token]
    
    if expired_sessions:
        logger.info(f"Cleaned up {len(expired_sessions)} expired chef centre sessions.")

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            await connection.send_text(message)

manager = ConnectionManager()

# Define data models
class VerificateurCreate(BaseModel):
    username: str
    password: str
    full_name: str
    aref: str
    dp: str
    ville: str
    lycee: str
    salle: str
    matiere: str
    cols: str
    roos: str
    Examen: str

class VerificateurUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    full_name: Optional[str] = None
    aref: Optional[str] = None
    dp: Optional[str] = None
    ville: Optional[str] = None
    lycee: Optional[str] = None
    salle: Optional[str] = None
    matiere: Optional[str] = None
    Examen: Optional[str] = None


class VerificateurLogin(BaseModel):
    username: str
    password: str

class AnalyseItem(BaseModel):
    date: str
    operateur: str
    type_communication: str

class AnalyseGeneraleCreate(BaseModel):
    id_verificateur: int
    salle: str
    matiere: str
    data: List[AnalyseItem]

class MobilityItemInitial(BaseModel):
    timestamp:str
    bond: float
    risk_level: str
    puissance: float
    cne:str


class AnalyseMobiliteInitialCreate(BaseModel):
    id_verificateur: int
    data: List[MobilityItemInitial]

class MobilityItemFinal(BaseModel):
    cne:str
    risk_level: int
    risk_status: str
    power: float

class AnalyseMobiliteFinalCreate(BaseModel):
    id_verificateur: int
    data: List[MobilityItemFinal]

class ConfigurationsUpdate(BaseModel):
    id_verificateur: int
    aref: str
    dp: str
    Ville: str
    Lycée: str
    Salle: str
    Matière: str
    cols: str
    roos: str
    Examen: str
# Add these helper functions to main.py

import hashlib

def hash_password(password: str) -> str:
    """Hash a password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def get_current_user(request: Request) -> dict:
    """Get the current user from the session"""
    session_token = request.cookies.get("session_token")
    if session_token and session_token in active_sessions:
        user_id = active_sessions[session_token]["user_id"]
        query = """
        SELECT u.id, u.full_name, u.username, u.profile_id, u.codeAref, p.profile_name, p.permissions
        FROM users u
        JOIN profiles p ON u.profile_id = p.profile_id
        WHERE u.id = %s AND u.is_active = TRUE
        """
        user = db.execute_query(query, (user_id,))
        if user:
            return user[0]
    return None

def get_user_aref_filter(request: Request) -> Optional[str]:
    """Get the AREF filter for the current user. Returns None for admin users, AREF code for AREF admins."""
    user = get_current_user(request)
    if not user:
        return None
    
    # If user is admin (profile_id=1), return None (no filter)
    if user["profile_id"] == 1:
        return None
    
    # If user is AREF admin (profile_id=4), return their AREF code
    if user["profile_id"] == 4:
        return user["codeAref"]
    
    # For other roles, return None (no filter)
    return None

# Add a route to get current user info
@app.get("/t3shield/api/current-user")
def current_user(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )
    return {
        "id": user["id"],
        "full_name": user["full_name"],
        "username": user["username"],
        "profile_id": user["profile_id"],
        "profile_name": user["profile_name"],
        "codeAref": user["codeAref"]
    }



@app.post("/t3shield/login")
def login(login_data: UserLogin, response: Response):
    # Hash the provided password
    hashed_password = login_data.password #hash_password(login_data.password)
    
    # Check credentials
    query = """
    SELECT u.id, u.full_name, u.username, u.profile_id, u.codeAref, p.profile_name, p.permissions
    FROM users u
    JOIN profiles p ON u.profile_id = p.profile_id
    WHERE u.username = %s AND u.password = %s AND u.is_active = TRUE
    """
    user = db.execute_query(query, (login_data.username, hashed_password))
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Nom d'utilisateur ou mot de passe incorrect"
        )
    
    # Update last login time
    update_query = "UPDATE users SET last_login = NOW() WHERE id = %s"
    db.execute_update(update_query, (user[0]["id"],))
    
    # Create a session token
    session_token = secrets.token_hex(16)
    
    # Store session information
    active_sessions[session_token] = {
        "user_id": user[0]["id"],
        "username": user[0]["username"],
        "profile_id": user[0]["profile_id"],
        "codeAref": user[0]["codeAref"],
        "permissions": user[0]["permissions"]
    }
    
    # Set session cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        max_age=3600 * 3,  # 24 hours
        samesite="lax"
    )
    
    # Return user information
    return {
        "id": user[0]["id"],
        "full_name": user[0]["full_name"],
        "username": user[0]["username"],
        "profile_id": user[0]["profile_id"],
        "profile_name": user[0]["profile_name"],
        "codeAref": user[0]["codeAref"]
    }

# Add logout endpoint
@app.post("/t3shield/api/logout")
def logout(response: Response, session_token: Optional[str] = Cookie(None)):
    if session_token and session_token in active_sessions:
        # Remove session
        active_sessions.pop(session_token)
    
    # Clear the cookie
    response.delete_cookie(key="session_token")
    
    return {"message": "Logged out successfully"}

# Add middleware to check authentication for protected routes
@app.middleware("http")
async def check_authentication(request: Request, call_next):
    # Paths that don't require authentication
    public_paths = [
        "/t3shield/login", 
        "/t3shield/api/login",
        "/t3shield/api/send_configurations",
        "/t3shield/api/download-configuration-files",
        "/t3shield/api/download-configuration",
        "/t3shield/api/download-aref-centres",
        "/t3shield/api/create_verificateur",
        "/t3shield/api/send_general_analysis",
        "/t3shield/api/send_mobility_analysis",
        "/verificateurs",
        "/t3shield/api/get_lycees",
        "/t3shield/api/get_villes",
        "/t3shield/api/get_dps",
        "/t3shield/api/get_arefs",
        "/t3shield/api/configuration/",
        "/t3shield/api/geojson/regions",
        "/t3shield/api/geojson/provinces",
        "/test/verified",
        "/api/analyses",
        "/api/mobility_analyses", 
        "/api/verified_analyses",
        "/api/statistics",
        "/health",

        "/login", 
        "/static/", 
        "/css/", 
        "/js/",
        "/favicon.ico"
    ]
    
    # Check if the path is public
    is_public = False
    for path in public_paths:
        if request.url.path.startswith(path):
            is_public = True
            break
    
    # Special handling for GeoJSON provinces endpoint with region ID
    if request.url.path.startswith("/t3shield/api/geojson/provinces/"):
        is_public = True
    
    # Special handling for chef center routes
    if request.url.path.startswith("/chef/centre/"):
        # Extract code_centre from URL
        path_parts = request.url.path.split("/")
        if len(path_parts) >= 4:
            code_centre = path_parts[3]
            
            # Public endpoints that don't require authentication
            public_chef_endpoints = [
                f"/chef/centre/{code_centre}",           # Main page (login page)
                f"/chef/centre/{code_centre}/login",     # Login endpoint
                f"/chef/centre/{code_centre}/logout"     # Logout endpoint
            ]
            
            # Check if this is a public endpoint
            if request.url.path in public_chef_endpoints:
                response = await call_next(request)
                return response
            
            # For protected routes, check authentication
            chef_session_token = request.cookies.get("chef_centre_session")
            
            # Clean up expired sessions periodically
            cleanup_expired_sessions()
            
            # Check if we have a valid session
            if chef_session_token and chef_session_token in chef_centre_sessions:
                session_data = chef_centre_sessions[chef_session_token]
                if session_data["code_centre"] == code_centre:
                    response = await call_next(request)
                    return response
            
            # No valid session - return 401 for API endpoints, unauthorized page for others
            if request.url.path.startswith(f"/chef/centre/{code_centre}/api/"):
                return JSONResponse(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    content={"detail": "Authentication required"}
                )
            else:
                # For dashboard and other pages, serve unauthorized page
                return FileResponse("front-end/chef-unauthorized.html")
        else:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={"detail": "Not found"}
            )
    
    if is_public:
        # Allow access to public paths
        response = await call_next(request)
        return response
    
    # Check for session token
    session_token = request.cookies.get("session_token")
    
    if not session_token or session_token not in active_sessions:
        # If API request, return 401
        if request.url.path.startswith("/t3shield/api/"):
            return JSONResponse(
                status_code=status.HTTP_401_UNAUTHORIZED,
                content={"detail": "Authentication required"}
            )
        # Otherwise redirect to login page
        return RedirectResponse(url="/login")
    
    # Continue with the request
    response = await call_next(request)
    return response

# Add login page route
@app.get("/login", include_in_schema=False)
def get_login_page(request: Request, session_token: Optional[str] = Cookie(None)):
    # If already logged in, redirect to dashboard
    if session_token and session_token in active_sessions:
        return RedirectResponse(url="/dashboard-national")
    
    return FileResponse("front-end/login.html")

def hash_password(password: str) -> str:
    """Hash password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

@app.get("/", include_in_schema=False)
def get_dashboard():
    return FileResponse("front-end/index.html")

@app.post("/t3shield/api/create_verificateur", status_code=status.HTTP_201_CREATED)
def create_verificateur(verificateur: VerificateurCreate):
    # Check if username already exists
    check_query = "SELECT * FROM verificateur WHERE username = %s"
    existing_user = db.execute_query(check_query, (verificateur.username,))
    
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already registered"
        )
    
    # Hash the password
    hashed_password = verificateur.password #hash_password(verificateur.password)
    
    # Insert new verificateur
    insert_query = """
    INSERT INTO verificateur (username, password, full_name, aref, dp, ville, lycee, salle, matiere, cols, roos)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    params = (
        verificateur.username,
        hashed_password,
        verificateur.full_name,
        verificateur.aref,
        verificateur.dp,
        verificateur.ville,
        verificateur.lycee,
        verificateur.salle,
        verificateur.matiere,
        verificateur.cols,
        verificateur.roos
    )
    
    if not db.execute_update(insert_query, params):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create verificateur"
        )
    
    # Get the created verificateur
    new_user = db.execute_query(check_query, (verificateur.username,))
    return new_user[0] if new_user else {"message": "Verificateur created successfully"}

def clean_province_name(province):
    """Clean province name by removing prefixes and hyphens."""
    patterns = [
        r'^Préfecture\s+d\'Arrond\.\s+',
        r'^d\'Arrond\.\s+',
        r'^Province:\s*',
        r'^Préfecture:\s*',
        r'^Province\s+',
        r'^Préfecture\s+'
    ]
    
    cleaned = province
    for pattern in patterns:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
    
    # Remove hyphens
    cleaned = cleaned.replace('-', '')
    
    return cleaned.strip()

@app.post("/t3shield/api/send_configurations", status_code=status.HTTP_200_OK)
def update_configurations(config_data: ConfigurationsUpdate):
    # Check if verificateur exists
    id_verificateur = config_data.id_verificateur
    check_query = "SELECT * FROM verificateur WHERE id_verificateur = %s"
    existing_user = db.execute_query(check_query, (id_verificateur,))
    
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Verificateur with ID {id_verificateur} not found"
        )
    
    # Update verificateur information
    update_query = """
    UPDATE verificateur 
    SET aref = %s, dp = %s, ville = %s, lycee = %s, salle = %s, matiere = %s, cols = %s, roos = %s, examen = %s
    WHERE id_verificateur = %s
    """
    
    # Extract code from aref string (e.g. "13 - AREF de Test" -> "13")
    aref_code = config_data.aref.split('-')[0].strip()
    
    # Get aref name from database using code
    aref_query = "SELECT name FROM aref WHERE codeAref = %s"
    aref_result = db.execute_query(aref_query, (aref_code,))

    aref = aref_result[0]["name"] if aref_result else config_data.aref
    dp = clean_province_name(config_data.dp)
    ville = dp  # ville takes same value as cleaned dp
    lycee = config_data.Lycée
    salle = config_data.Salle
    matiere = config_data.Matière
    cols = config_data.cols
    roos = config_data.roos
    examen = config_data.Examen
    params = (
        aref,
        dp,
        ville,
        lycee,
        salle,
        matiere,
        cols,
        roos,
        examen,
        id_verificateur
    )
    
    if not db.execute_update(update_query, params):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update verificateur configurations"
        )
    
    # Handle chef_center activation logic
    try:
        # First, deactivate any existing active center for this verificateur
        deactivate_query = """
        UPDATE chef_center 
        SET active = 0 
        WHERE id_verificateur = %s AND active = 1
        """
        db.execute_update(deactivate_query, (id_verificateur,))
        # get the codeCentre from lycee
        codeCentre_query = """
        SELECT codeCentre FROM lycee WHERE name_fr = %s
        """
        codeCentre = db.execute_query(codeCentre_query, (lycee,))
        codeCentre = codeCentre[0]['codeCentre']

        # Then activate the center for this lycee (assuming codeCentre matches lycee name)
        activate_query = """
        UPDATE chef_center 
        SET active = 1, id_verificateur = %s 
        WHERE codeCentre = %s
        """
        result = db.execute_update(activate_query, (id_verificateur, codeCentre))
        
        if not result:
            logger.warning(f"No chef_center found for lycee: {lycee}")
            
    except Exception as e:
        logger.error(f"Error updating chef_center: {str(e)}")
        # Don't fail the whole request if chef_center update fails
    
    # Get the updated verificateur
    updated_user = db.execute_query(check_query, (config_data.id_verificateur,))
    return updated_user[0] if updated_user else {"message": "Configurations updated successfully"}



@app.post("/t3shield/api/login")
def login(login_data: VerificateurLogin):
    # Hash the provided password
    hashed_password = login_data.password #hash_password(login_data.password)
    
    # Check credentials
    query = "SELECT id_verificateur, full_name, aref, dp, ville, lycee, salle, matiere, cols, roos, examen FROM verificateur WHERE username = %s AND password = %s"
    user = db.execute_query(query, (login_data.username, hashed_password))
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Nom d'utilisateur ou mot de passe incorrect"
        )
    
    return user[0]

@app.get("/verificateurs", response_model=List[Dict[str, Any]])
def get_all_verificateurs():
    query = "SELECT id_verificateur, username, full_name, aref, dp, ville, lycee, salle, matiere, cols, roos FROM verificateur"
    users = db.execute_query(query)
    
    if users is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve verificateurs"
        )
    
    return users

@app.get("/verificateurs/{verificateur_id}")
def get_verificateur(verificateur_id: int):
    query = "SELECT id_verificateur, username, full_name, aref, dp, ville, lycee, salle, matiere, cols, roos FROM verificateur WHERE id_verificateur = %s"
    user = db.execute_query(query, (verificateur_id,))
    
    if user is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve verificateur"
        )
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Verificateur with ID {verificateur_id} not found"
        )
    return user[0]

@app.post("/t3shield/api/send_general_analysis", status_code=status.HTTP_201_CREATED)
async def send_general_analysis(analyse_data: AnalyseGeneraleCreate):
    id_verificateur = analyse_data.id_verificateur
    # Check if verificateur exists
    check_query = "SELECT * FROM verificateur WHERE id_verificateur = %s"
    existing_user = db.execute_query(check_query, (id_verificateur,))
    
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Verificateur with ID {id_verificateur} not found"
        )
    
    # Get verificateur details
    verificateur = existing_user[0]
    
    # Get the latest batch number for this location and subject
    batch_query = """
    SELECT MAX(batch) as latest_batch 
    FROM analyse_generale 
    WHERE aref = %s AND dp = %s AND ville = %s AND lycee = %s AND salle = %s AND matiere = %s
    """
    aref = verificateur["aref"]
    dp = verificateur["dp"]
    ville = verificateur["ville"]
    lycee = verificateur["lycee"]
    salle = verificateur["salle"]
    matiere = verificateur["matiere"]
    general_analysis_data = analyse_data.data
    general_analysis_count = len(general_analysis_data)
    cols = int(verificateur["cols"])
    roos = int(verificateur["roos"])
    
    batch_params = (
        aref,
        dp,
        ville,
        lycee,
        salle,
        matiere
    )
    
    latest_batch = db.execute_query(batch_query, batch_params)

    # Determine the new batch number
    new_batch = 1
    if latest_batch and latest_batch[0]['latest_batch'] is not None:
        new_batch = latest_batch[0]['latest_batch'] + 1

    # Insert query for analyse_generale - updated to include batch
    insert_query = """
    INSERT INTO analyse_generale (
        date, operateur, type_communication, id_verificateur, 
        salle, matiere, aref, dp, ville, lycee, batch, cols, roos
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    inserted_records = []
    
    # Process each item in the data list
    for item in general_analysis_data:

        if (item.operateur == "OperateurI"):
            item.operateur = "INWI"
        elif (item.operateur == "OperateurM"):
            item.operateur = "IAM"
        elif (item.operateur == "OperateurO"):
            item.operateur = "ORANGE"
        if(item.type_communication == "Protocole900"):
            item.type_communication = "GSM"
        params = (
            item.date,
            item.operateur,
            item.type_communication,
            id_verificateur,
            salle,
            matiere,
            aref,
            dp,
            ville,
            lycee,
            new_batch,
            cols,
            roos
        )
        
        if db.execute_update(insert_query, params):
            # Create a record for WebSocket broadcast
            analysis_record = {
                "type": "analyse_generale",
                "date": item.date,
                "operateur": item.operateur,
                "type_communication": item.type_communication,
                "aref": aref,
                "dp": dp,
                "ville": ville,
                "lycee": lycee,
                "salle": salle,
                "matiere": matiere,
                "verificateur_name": verificateur["full_name"],
                "batch": new_batch  # Include batch in the response
            }
            inserted_records.append(analysis_record)
    
    # Broadcast the new analyses to all connected WebSocket clients
    if inserted_records:

        await manager.broadcast(json.dumps({
            "event": "new_analyses",
            "data": inserted_records
        }))
    
    # Handle cheat rate insertion for empty data lists (no suspicious activity detected)
    cheat_rate_updated = False
    if len(analyse_data.data) == 0:
        try:

            # Get codeCentre from lycee table
            codeCentre_query = "SELECT codeCentre FROM lycee WHERE name_fr = %s"
            codeCentre_result = db.execute_query(codeCentre_query, (lycee,))
            
            if codeCentre_result:
                codeCentre = codeCentre_result[0]['codeCentre']
                examen = verificateur.get("examen", "Session Normale")  # Get examen from verificateur or default
                current_date_time = get_morocco_time_str()

                # MODIFIED: Always insert new record to preserve historical data
                # No longer check for existing records or update them
                insert_cheat_rate_query = """
                INSERT INTO cheat_rate (codeCentre, salle, matiere, examen, date, nbr_etudiant, nbr_detection)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                print(f"cols: {cols} * roos: {roos} = {cols*roos}")
                insert_cheat_rate_params = (codeCentre, salle, matiere, examen, current_date_time, (cols*roos), 0)
                
                db.execute_update(insert_cheat_rate_query, insert_cheat_rate_params)

                cheat_rate_updated = True
                
                # Broadcast cheat rate update for real-time stats
                await manager.broadcast(json.dumps({
                    "event": "cheat_rate_update",
                    "type": "clean_session",
                    "data": {
                        "codeCentre": codeCentre,
                        "aref": aref,
                        "dp": dp,
                        "ville": ville,
                        "lycee": lycee,
                        "salle": salle,
                        "matiere": matiere,
                        "examen": examen,
                        "date": current_date_time,
                        "nbr_etudiant": 20,
                        "nbr_detection": 0,
                        "verificateur_name": verificateur["full_name"],
                        "message": "Clean session recorded - no suspicious activity detected"
                    }
                }))
                
        except Exception as e:

            logger.error(f"Error handling general analysis : {e}")
    
    # Return success message
    return {
        "message": "Analysis data successfully recorded", 
        "count": len(analyse_data.data), 
        "batch": new_batch,
        "cheat_rate_updated": cheat_rate_updated,
        "is_clean_session": len(analyse_data.data) == 0
    }

@app.post("/t3shield/api/send_mobility_analysis", status_code=status.HTTP_201_CREATED)
async def send_mobility_analysis(mobility_data: AnalyseMobiliteInitialCreate):
    # Check if verificateur exists
    id_verificateur = mobility_data.id_verificateur
    check_query = "SELECT * FROM verificateur WHERE id_verificateur = %s"
    existing_user = db.execute_query(check_query, (id_verificateur,))
    
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Verificateur with ID {id_verificateur} not found"
        )
    
    # Get verificateur details
    verificateur = existing_user[0]
    

    # Insert query for analyse_mobilite - NOTE: Using 'date' instead of 'timestamp'
    insert_query = """
    INSERT INTO analyse_mobilite (
        id_verificateur, salle, matiere, cne, timestamp, operator, risk_level, puissance, bond,
        aref, dp, ville, lycee, cols, roos
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    aref = verificateur["aref"]
    dp = verificateur["dp"]
    ville = verificateur["ville"]
    lycee = verificateur["lycee"]
    salle = verificateur["salle"]
    matiere = verificateur["matiere"]
    cols = int(verificateur["cols"])
    roos = int(verificateur["roos"])
    initial_mobility_analysis_data = mobility_data.data
    initial_mobility_analysis_count = len(initial_mobility_analysis_data)
    current_date_time = get_morocco_time_str()
    # Insert initial records first
    for item in initial_mobility_analysis_data:
        params = (
            id_verificateur,
            salle,
            matiere,
            item.cne,
            item.timestamp,  # Using timestamp from item but inserting into 'date' column
            "ORANGE",
            item.risk_level,
            item.puissance,
            item.bond,
            aref,
            dp,
            ville,
            lycee,
            cols,
            roos
        )
        
        db.execute_update(insert_query, params)
    
    return {"message": "Mobility analysis data successfully recorded", "count": initial_mobility_analysis_count}

@app.post("/t3shield/api/send_mobility_analysis_final", status_code=status.HTTP_201_CREATED)
async def send_mobility_analysis_final(mobility_data_final: AnalyseMobiliteFinalCreate):
    # Check if verificateur exists
    id_verificateur = mobility_data_final.id_verificateur

    check_query = "SELECT * FROM verificateur WHERE id_verificateur = %s"
    existing_user = db.execute_query(check_query, (id_verificateur,))
    
    if not existing_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Verificateur with ID {id_verificateur} not found"
        )
    
    # Get verificateur details
    verificateur = existing_user[0]
    
    # Get codeCentre from lycee table
    codeCentre_query = "SELECT codeCentre FROM lycee WHERE name_fr = %s"
    codeCentre_result = db.execute_query(codeCentre_query, (verificateur["lycee"],))
    
    if not codeCentre_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Lycee {verificateur['lycee']} not found in lycee table"
        )
    
    codeCentre = codeCentre_result[0]['codeCentre']
    
    # Insert query for analyse_mobility_final - updated to include batch
    insert_query = """
    INSERT INTO analyse_mobility_final (
        id_verificateur, salle, matiere, cne, timestamp, operator,
        aref, dp, ville, lycee, batch, cols, roos, risk_status, power
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    
    inserted_records = []
    
    # Get current date and time in GMT+1 (Morocco timezone)
    current_date = get_morocco_date_str()
    current_date_time = get_morocco_time_str()
    aref = verificateur["aref"]
    dp = verificateur["dp"]
    ville = verificateur["ville"]
    lycee = verificateur["lycee"]
    salle = verificateur["salle"]
    matiere = verificateur["matiere"]
    examen = verificateur["examen"]
    cols = int(verificateur["cols"])
    roos = int(verificateur["roos"])
    final_mobility_analysis_data = mobility_data_final.data
    final_mobility_analysis_count = len(final_mobility_analysis_data)

    # Calculate cheat rate statistics
    total_students = len(final_mobility_analysis_data)
    total_detections = sum(1 for item in final_mobility_analysis_data if item.risk_level == 1)

    # Process each item in the data list - using dictionary access
    for item in final_mobility_analysis_data:
        cne = item.cne
        risk_status = item.risk_status
        power = item.power
        timestamp = current_date_time
        operator = "ORANGE"
        if item.risk_level == 1:
            # Get the latest batch number for this location, subject, and CNE for today only
            batch_query = """
            SELECT MAX(batch) as latest_batch 
            FROM analyse_mobility_final 
            WHERE aref = %s AND dp = %s AND ville = %s AND lycee = %s AND salle = %s AND matiere = %s AND cne = %s
            AND DATE(timestamp) = %s
            """
            
            batch_params = (
                aref,
                dp,
                ville,
                lycee,
                salle,
                matiere,
                cne,  
                current_date  # Added current date parameter
            )

            latest_batch = db.execute_query(batch_query, batch_params)
            
            # Determine the new batch number
            new_batch = 1
            if latest_batch and latest_batch[0]['latest_batch'] is not None:
                new_batch = latest_batch[0]['latest_batch'] + 1

            params = (
                id_verificateur,
                salle,
                matiere,
                item.cne,
                timestamp,  # Using timestamp from item but inserting into 'date' column
                operator,
                aref,
                dp,
                ville,
                lycee,
                new_batch,
                cols,
                roos,
                risk_status,
                power
            )

            if db.execute_update(insert_query, params):
                # Create a record for WebSocket broadcast
                mobility_analysis_record = {
                    "type": "analyse_mobilite",
                    "timestamp": timestamp,  # Using "date" as the key in response
                    "operator": operator,
                    "aref": aref,
                    "dp": dp,
                    "ville": ville,
                    "lycee": lycee,
                    "salle": salle,
                    "matiere": matiere,
                    "cne": cne,
                    "verificateur_name": verificateur["full_name"] if "full_name" in verificateur else "Unknown",
                    "batch": new_batch  # Include batch in the response
                }
                inserted_records.append(mobility_analysis_record)
    
    # Handle cheat rate insertion/update
    cheat_rate_updated = False
    try:
        # MODIFIED: Always insert new record to preserve historical data
        # No longer check for existing records or update them
        insert_cheat_rate_query = """
        INSERT INTO cheat_rate (codeCentre, salle, matiere, examen, date, nbr_etudiant, nbr_detection)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        insert_cheat_rate_params = (codeCentre, salle, matiere, examen, current_date_time, (cols*roos), total_detections)
        db.execute_update(insert_cheat_rate_query, insert_cheat_rate_params)

        cheat_rate_updated = True
    
    except Exception as e:

        logger.error(f"Error handling mobility analysis final: {e}")
    
    # Always broadcast if we have inserted records OR if cheat rate was updated
    if inserted_records or cheat_rate_updated:
        await manager.broadcast(json.dumps({
            "event": "new_mobility_analysis",
            "data": inserted_records,
            "cheat_rate_updated": cheat_rate_updated,
            "total_students": total_students,
            "total_detections": total_detections
        }))
    
    # Send dedicated cheat rate update for real-time stats dashboard
    if cheat_rate_updated:
        await manager.broadcast(json.dumps({
            "event": "cheat_rate_update",
            "type": "detection_session",
            "data": {
                "codeCentre": codeCentre,
                "aref": aref,
                "dp": dp,
                "ville": ville,
                "lycee": lycee,
                "salle": salle,
                "matiere": matiere,
                "examen": examen,
                "date": current_date_time,
                "nbr_etudiant": total_students,
                "nbr_detection": total_detections,
                "verificateur_name": verificateur["full_name"],
                "message": f"Detection session completed - {total_detections} detections out of {total_students} students"
            }
        }))
    return {"message": "Mobility analysis data successfully recorded", "count": final_mobility_analysis_count}


@app.get("/api/analyses")
def get_analyses(start_time: Optional[str] = None, end_time: Optional[str] = None, request: Request = None):
    """Get analyses with optional time filtering - only latest batch per location per day"""
    
    # Default to today 8am-6pm if no time range provided (GMT+1)
    if not start_time and not end_time:
        today = get_morocco_date_str()
        start_time = f"{today} 08:00:00"
        end_time = f"{today} 18:00:00"
    
    # Get user AREF filter
    user_aref = get_user_aref_filter(request) if request else None
    
    # Log the received parameters for debugging

    # Build the query with batch filtering - only show latest batch per location per day
    # FIXED: Filter based on analysis location data (ag.aref) instead of verificateur's current AREF
    query_parts = [
        "SELECT ag.*, v.full_name as verificateur_name",
        "FROM analyse_generale ag",
        "JOIN verificateur v ON ag.id_verificateur = v.id_verificateur",
        "WHERE ag.batch = (",
        "    SELECT MAX(ag2.batch)",
        "    FROM analyse_generale ag2",
        "    WHERE ag2.aref = ag.aref AND ag2.dp = ag.dp AND ag2.ville = ag.ville",
        "    AND ag2.lycee = ag.lycee AND ag2.salle = ag.salle AND ag2.matiere = ag.matiere",
        "    AND DATE(ag2.date) = DATE(ag.date)",
        ")"
    ]
    params = []
    
    # Add AREF filter if user is AREF admin - filter by analysis location, not verificateur
    if user_aref:
        query_parts.append("AND ag.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)")
        params.append(user_aref)
    
    # Add time filters if provided
    if start_time:
        query_parts.append("AND ag.date >= %s")
        params.append(start_time)
    
    if end_time:
        query_parts.append("AND ag.date <= %s")
        params.append(end_time)
    
    # Order by date
    query_parts.append("ORDER BY ag.date DESC")
    
    # Execute the query
    query = " ".join(query_parts)

    analyses = db.execute_query(query, tuple(params))
    
    if analyses is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve analyses"
        )
    
    # Transform the data for the dashboard
    result = []
    for analysis in analyses:
        result.append({
            "id": analysis["id_analyse"],
            "type": "analyse_generale",
            "date": analysis["date"],
            "operateur": analysis["operateur"],
            "type_communication": analysis["type_communication"],
            "aref": analysis["aref"],
            "dp": analysis["dp"],
            "ville": analysis["ville"],
            "lycee": analysis["lycee"],
            "salle": analysis["salle"],
            "matiere": analysis["matiere"],
            "verificateur_name": analysis["verificateur_name"],
            "batch": analysis["batch"] if "batch" in analysis else 1
        })
    
    # Log the number of results

    return result

@app.get("/api/mobility_analyses")
def get_mobility_analyses(start_time: Optional[str] = None, end_time: Optional[str] = None, request: Request = None):
    """Get analyses with optional time filtering"""
    
    # Default to today 8am-6pm if no time range provided (GMT+1)
    if not start_time and not end_time:
        today = get_morocco_date_str()
        start_time = f"{today} 08:00:00"
        end_time = f"{today} 18:00:00"
    
    # Get user AREF filter
    user_aref = get_user_aref_filter(request) if request else None
    
    # Log the received parameters for debugging

    # Build the query - IMPORTANT: Exclude verified analyses AND only show latest batch per location+CNE PER DAY
    # FIXED: Filter based on analysis location data (ag.aref) instead of verificateur's current AREF
    query_parts = [
        "SELECT ag.*, v.full_name as verificateur_name",
        "FROM analyse_mobility_final ag",
        "JOIN verificateur v ON ag.id_verificateur = v.id_verificateur",
        "LEFT JOIN analyse_verifier av ON ag.id_analyse = av.id_analyse_mobilite_final",
        "WHERE av.id_analyse_mobilite_final IS NULL",  # Only get unverified analyses
        "AND ag.batch = (",
        "    SELECT MAX(amf2.batch)",
        "    FROM analyse_mobility_final amf2",
        "    WHERE amf2.aref = ag.aref AND amf2.dp = ag.dp AND amf2.ville = ag.ville",
        "    AND amf2.lycee = ag.lycee AND amf2.salle = ag.salle AND amf2.matiere = ag.matiere",
        "    AND amf2.cne = ag.cne",
        "    AND DATE(amf2.timestamp) = DATE(ag.timestamp)",
        ")"
    ]
    params = []
    
    # Add AREF filter if user is AREF admin - filter by analysis location, not verificateur
    if user_aref:
        query_parts.append("AND ag.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)")
        params.append(user_aref)
    
    # Add time filters if provided - use simple string comparison instead of STR_TO_DATE
    if start_time:
        query_parts.append("AND ag.timestamp >= %s")
        params.append(start_time)
    
    if end_time:
        query_parts.append("AND ag.timestamp <= %s")
        params.append(end_time)
    
    # Order by date - use simple ordering
    query_parts.append("ORDER BY ag.timestamp DESC")
    
    # Execute the query
    query = " ".join(query_parts)

    analyses = db.execute_query(query, tuple(params))
    
    if analyses is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve analyses"
        )
    
    # Transform the data for the dashboard
    result = []
    for analysis in analyses:
        result.append({
            "id": analysis["id_analyse"],
            "type": "analyse_mobilite",
            "date": analysis["timestamp"],  # Use timestamp as date for consistency
            "operateur": analysis["operator"],
            "aref": analysis["aref"],
            "dp": analysis["dp"],
            "ville": analysis["ville"],
            "lycee": analysis["lycee"],
            "salle": analysis["salle"],
            "matiere": analysis["matiere"],
            "verificateur_name": analysis["verificateur_name"],
            "cne": analysis["cne"],  # Include CNE for mobility analyses
            "batch": analysis["batch"] if "batch" in analysis else 1  # Include batch, default to 1 if not present
        })
    
    # Log the number of results

    return result


@app.get("/api/verified_analyses")
def get_verified_analyses(start_time: Optional[str] = None, end_time: Optional[str] = None, request: Request = None):
    """Get the latest verified/denied analyses with optional time filtering"""
    
    if not start_time and not end_time:
        today = get_morocco_date_str()
        start_time = f"{today} 08:00:00"
        end_time = f"{today} 18:00:00"
    
    # Get user AREF filter
    user_aref = get_user_aref_filter(request) if request else None

            # Corrected query with aliases to match frontend expectations
    # FIXED: Filter based on analysis location data (am.aref) instead of verificateur's current AREF
    query = """
        WITH RankedVerifications AS (
            SELECT 
                av.id_analyse_verifier as id, 
                av.id_analyse_mobilite_final as id_analyse_mobilite,
                av.action,
                av.timestamp as verification_timestamp,
                am.timestamp as date,
                am.operator as operateur,
                am.aref, am.dp, am.ville, am.lycee,
                am.salle, am.matiere, am.cne, am.batch,
                v.full_name as verificateur_name,
                ROW_NUMBER() OVER(PARTITION BY av.id_analyse_mobilite_final ORDER BY av.id_analyse_verifier DESC) as rn
            FROM analyse_verifier av
            JOIN analyse_mobility_final am ON av.id_analyse_mobilite_final = am.id_analyse
            JOIN verificateur v ON am.id_verificateur = v.id_verificateur
            WHERE am.timestamp BETWEEN %s AND %s
    """
    
    params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin - filter by analysis location, not verificateur
    if user_aref:
        query += " AND am.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        params.append(user_aref)
    
    query += ") SELECT * FROM RankedVerifications WHERE rn = 1"
    
    verified_analyses = db.execute_query(query, tuple(params))
    
    if verified_analyses is None:
        verified_analyses = []
    
    # Transform the data to match expected format
    result = []
    for analysis in verified_analyses:
        result.append({
            "id": analysis["id"],
            "id_analyse_mobilite": analysis["id_analyse_mobilite"],
            "action": analysis["action"],
            "verification_timestamp": analysis["verification_timestamp"],
            "date": analysis["date"],
            "operateur": analysis["operateur"],
            "aref": analysis["aref"],
            "dp": analysis["dp"],
            "ville": analysis["ville"],
            "lycee": analysis["lycee"],
            "salle": analysis["salle"],
            "matiere": analysis["matiere"],
            "cne": analysis["cne"],
            "batch": analysis["batch"],
            "verificateur_name": analysis["verificateur_name"]
        })

    return result

@app.get("/api/statistics")
def get_statistics(
    start_time: Optional[str] = None, 
    end_time: Optional[str] = None,
    request: Request = None
):
    """Get statistics data for the dashboard with time filter only"""
    
    # Get current date for default time range (GMT+1)
    today = get_morocco_date_str()
    start_time = start_time or f"{today} 08:00:00"
    end_time = end_time or f"{today} 18:00:00"
    
    # Get user AREF filter
    user_aref = get_user_aref_filter(request) if request else None
    
    # Log the received parameters for debugging

    # Build base query parts for general analyses - only show latest batch per location per day
    # FIXED: Filter based on analysis location data (ag.aref) instead of verificateur's current AREF
    general_query = """
    SELECT ag.*, v.full_name as verificateur_name
    FROM analyse_generale ag
    JOIN verificateur v ON ag.id_verificateur = v.id_verificateur
    WHERE ag.date BETWEEN %s AND %s
    AND ag.batch = (
        SELECT MAX(ag2.batch)
        FROM analyse_generale ag2
        WHERE ag2.aref = ag.aref AND ag2.dp = ag.dp AND ag2.ville = ag.ville
        AND ag2.lycee = ag.lycee AND ag2.salle = ag.salle AND ag2.matiere = ag.matiere
        AND DATE(ag2.date) = DATE(ag.date)
    )
    """
    
    general_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin - filter by analysis location, not verificateur
    if user_aref:
        general_query += " AND ag.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        general_params.append(user_aref)
    
    # Execute the query for general analyses
    general_analyses = db.execute_query(general_query, tuple(general_params))
    
    if general_analyses is None:
        general_analyses = []
    
    # Build query for mobility analyses - exclude verified ones and only show latest batch per location+CNE per day
    # FIXED: Filter based on analysis location data (ag.aref) instead of verificateur's current AREF
    mobility_query = """
    SELECT ag.*, v.full_name as verificateur_name
    FROM analyse_mobility_final ag
    JOIN verificateur v ON ag.id_verificateur = v.id_verificateur
    LEFT JOIN analyse_verifier av ON ag.id_analyse = av.id_analyse_mobilite_final
    WHERE av.id_analyse_mobilite_final IS NULL
    AND ag.batch = (
        SELECT MAX(amf2.batch)
        FROM analyse_mobility_final amf2
        WHERE amf2.aref = ag.aref AND amf2.dp = ag.dp AND amf2.ville = ag.ville
        AND amf2.lycee = ag.lycee AND amf2.salle = ag.salle AND amf2.matiere = ag.matiere
        AND amf2.cne = ag.cne
        AND DATE(amf2.timestamp) = DATE(ag.timestamp)
    )
    AND ag.timestamp BETWEEN %s AND %s
    """
    
    mobility_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin - filter by analysis location, not verificateur
    if user_aref:
        mobility_query += " AND ag.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        mobility_params.append(user_aref)
    
    # Execute the query for mobility analyses
    mobility_analyses = db.execute_query(mobility_query, tuple(mobility_params))
    
    if mobility_analyses is None:
        mobility_analyses = []
    
    # Fetch verified analyses (confirmed and denied)
    # FIXED: Filter based on analysis location data (am.aref) instead of verificateur's current AREF
    verified_query = """
    WITH RankedVerifications AS (
        SELECT 
            av.id_analyse_verifier as id, 
            av.id_analyse_mobilite_final as id_analyse_mobilite,
            av.action,
            av.timestamp as verification_timestamp,
            am.timestamp as date,
            am.operator as operateur,
            am.aref, am.dp, am.ville, am.lycee,
            am.salle, am.matiere, am.cne, am.batch,
            v.full_name as verificateur_name,
            ROW_NUMBER() OVER(PARTITION BY av.id_analyse_mobilite_final ORDER BY av.id_analyse_verifier DESC) as rn
        FROM analyse_verifier av
        JOIN analyse_mobility_final am ON av.id_analyse_mobilite_final = am.id_analyse
        JOIN verificateur v ON am.id_verificateur = v.id_verificateur
        WHERE am.timestamp BETWEEN %s AND %s
    """
    
    verified_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin - filter by analysis location, not verificateur
    if user_aref:
        verified_query += " AND am.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        verified_params.append(user_aref)
    
    verified_query += ") SELECT * FROM RankedVerifications WHERE rn = 1"
    
    verified_analyses = db.execute_query(verified_query, tuple(verified_params))
    
    if verified_analyses is None:
        verified_analyses = []
    
    # Since we now filter at SQL level, no need for batch filtering here
    # 1. Get total counts
    total_general = len(general_analyses)
    total_mobility = len(mobility_analyses)
    total_verified = len(verified_analyses)
    
    # Count verified vs denied
    verified_count = 0
    denied_count = 0
    for analysis in verified_analyses:
        if analysis["action"] == "confirm":
            verified_count += 1
        else:
            denied_count += 1
    
    # Combine both datasets for aggregation
    all_analyses = []
    
    # Add general analyses
    for analysis in general_analyses:
        all_analyses.append({
            "date": analysis["date"],
            "aref": analysis["aref"],
            "dp": analysis["dp"],
            "ville": analysis["ville"],
            "lycee": analysis["lycee"],
            "salle": analysis["salle"],
            "matiere": analysis["matiere"],
            "operateur": analysis["operateur"],
            "type_communication": analysis["type_communication"],
            "verificateur_name": analysis["verificateur_name"],
            "source": "general"
        })
    
    # Add mobility analyses
    for analysis in mobility_analyses:
        all_analyses.append({
            "date": analysis["timestamp"],  # Note: mobility uses 'timestamp' field
            "aref": analysis["aref"],
            "dp": analysis["dp"],
            "ville": analysis["ville"],
            "lycee": analysis["lycee"],
            "salle": analysis["salle"],
            "matiere": analysis["matiere"],
            "operateur": analysis["operator"],
            "type_communication": "data",  # Mobility analyses are typically data-based
            "verificateur_name": analysis["verificateur_name"],
            "source": "mobility"
        })
    
    # 2. Get counts by region (AREF)
    region_counts = {}
    for analysis in all_analyses:
        aref = analysis["aref"]
        if aref not in region_counts:
            region_counts[aref] = 0
        region_counts[aref] += 1
    
    regions = [{"name": aref, "count": count} for aref, count in region_counts.items()]
    regions.sort(key=lambda x: x["count"], reverse=True)
    
    # 3. Get counts by time of day
    time_of_day = {
        "morning": 0,
        "afternoon": 0,
        "evening": 0,
        "night": 0
    }
    
    for analysis in all_analyses:
        try:
            hour = datetime.strptime(analysis["date"], '%Y-%m-%d %H:%M:%S').hour
            if 6 <= hour < 12:
                time_of_day["morning"] += 1
            elif 12 <= hour < 18:
                time_of_day["afternoon"] += 1
            elif 18 <= hour < 24:
                time_of_day["evening"] += 1
            else:
                time_of_day["night"] += 1
        except Exception as e:
            logger.error(f"Error getting time of day: {e}")

    # 4. Get counts by communication type
    vocal_count = 0
    data_count = 0
    
    for analysis in all_analyses:
        if analysis["type_communication"].lower() in ["gsm", "vocal", "appel vocal"]:
            vocal_count += 1
        else:
            data_count += 1
    
    # 5. Get hourly distribution (dynamic based on user's time range)
    # Calculate the actual time range from the user's filter
    start_datetime = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    end_datetime = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
    
    # Get the min and max hours from the actual data
    min_hour = 23  # Start with max hour
    max_hour = 0   # Start with min hour
    
    for analysis in all_analyses:
        try:
            hour = datetime.strptime(analysis["date"], '%Y-%m-%d %H:%M:%S').hour
            min_hour = min(min_hour, hour)
            max_hour = max(max_hour, hour)
        except Exception as e:
            logger.error(f"Error getting hourly distribution: {e}")

    # If no data found, use the filter range
    if min_hour == 23 and max_hour == 0:
        min_hour = start_datetime.hour
        max_hour = end_datetime.hour
    
    # Ensure we have a reasonable range (at least 1 hour)
    if min_hour > max_hour:
        min_hour = 0
        max_hour = 23
    
    # 5.1. Get hourly distribution by region (AREF) - CUMULATIVE
    hourly_by_region = {}
    
    # Initialize hourly counts for each region
    for analysis in all_analyses:
        aref = analysis["aref"]
        if aref not in hourly_by_region:
            hourly_by_region[aref] = {}
            # Always use the full filter range for hours, not just the data range
            for hour in range(start_datetime.hour, end_datetime.hour + 1):
                hourly_by_region[aref][hour] = 0
    
    # Calculate cumulative counts by region
    for aref in hourly_by_region:
        # Get analyses for this region
        region_analyses = [a for a in all_analyses if a["aref"] == aref]
        sorted_region_analyses = sorted(region_analyses, key=lambda x: datetime.strptime(x["date"], '%Y-%m-%d %H:%M:%S'))
        
        # Calculate cumulative counts for this region
        current_cumulative = 0
        for analysis in sorted_region_analyses:
            try:
                hour = datetime.strptime(analysis["date"], '%Y-%m-%d %H:%M:%S').hour
                if start_datetime.hour <= hour <= end_datetime.hour:
                    current_cumulative += 1
                    # Update all hours from this hour onwards
                    for h in range(hour, end_datetime.hour + 1):
                        hourly_by_region[aref][h] = current_cumulative
            except Exception as e:
                logger.error(f"Error calculating cumulative counts by region: {e}")

        # Handle verified analyses for this region
        region_verified = [v for v in verified_analyses if v["aref"] == aref]
        sorted_region_verified = sorted(region_verified, key=lambda x: datetime.strptime(x["verification_timestamp"], '%Y-%m-%d %H:%M:%S'))
        
        for verification in sorted_region_verified:
            try:
                hour = datetime.strptime(verification["verification_timestamp"], '%Y-%m-%d %H:%M:%S').hour
                if start_datetime.hour <= hour <= end_datetime.hour:
                    # Subtract from all hours from this verification hour onwards
                    for h in range(hour, end_datetime.hour + 1):
                        hourly_by_region[aref][h] = max(0, hourly_by_region[aref][h] - 1)
            except Exception as e:
                logger.error(f"Error handling verified analyses for this region: {e}")

    # Format hourly data by region
    hourly_distribution_by_region = []
    for aref, hourly_counts in hourly_by_region.items():
        region_data = {
            'region': aref,
            'data': []
        }
        
        # Find the latest non-zero value to carry forward
        latest_value = 0
        for hour in range(start_datetime.hour, end_datetime.hour + 1):
            current_value = hourly_counts.get(hour, 0)
            if current_value > 0:
                latest_value = current_value
            region_data['data'].append({
                'hour': f"{hour:02d}:00",
                'count': latest_value  # Use latest known value for all hours
            })
        hourly_distribution_by_region.append(region_data)
    
    # 5.2. Get overall hourly distribution (CUMULATIVE - for backward compatibility)
    # Sort all analyses by timestamp for cumulative calculation
    sorted_analyses = sorted(all_analyses, key=lambda x: datetime.strptime(x["date"], '%Y-%m-%d %H:%M:%S'))
    
    # Initialize cumulative counts for each hour in the filter range
    cumulative_counts = {}
    # Always use the full filter range for hours, not just the data range
    for hour in range(start_datetime.hour, end_datetime.hour + 1):
        cumulative_counts[hour] = 0
    
    # Calculate cumulative counts
    current_cumulative = 0
    for analysis in sorted_analyses:
        try:
            hour = datetime.strptime(analysis["date"], '%Y-%m-%d %H:%M:%S').hour
            if start_datetime.hour <= hour <= end_datetime.hour:
                current_cumulative += 1
                # Update all hours from this hour onwards
                for h in range(hour, end_datetime.hour + 1):
                    cumulative_counts[h] = current_cumulative
        except Exception as e:
            logger.error(f"Error calculating cumulative counts: {e}")

    # Handle verified analyses by subtracting them from cumulative counts
    # Sort verified analyses by verification timestamp
    sorted_verified = sorted(verified_analyses, key=lambda x: datetime.strptime(x["verification_timestamp"], '%Y-%m-%d %H:%M:%S'))
    
    verified_cumulative = 0
    for verification in sorted_verified:
        try:
            hour = datetime.strptime(verification["verification_timestamp"], '%Y-%m-%d %H:%M:%S').hour
            if start_datetime.hour <= hour <= end_datetime.hour:
                verified_cumulative += 1
                # Subtract from all hours from this verification hour onwards
                for h in range(hour, end_datetime.hour + 1):
                    cumulative_counts[h] = max(0, cumulative_counts[h] - 1)
        except Exception as e:
            logger.error(f"Error handling verified analyses: {e}")

    formatted_hourly = []
    latest_cumulative_count = 0  # Track the latest known cumulative count
    
    for hour in range(start_datetime.hour, end_datetime.hour + 1):
        current_count = cumulative_counts[hour]
        if current_count > 0:
            latest_cumulative_count = current_count
        formatted_hourly.append({
            'hour': f"{hour:02d}:00",
            'count': latest_cumulative_count  # Use latest known count for all hours
        })
    
    # 5.5. Get detailed temporal distribution (day + hour) for multi-day ranges
    # Calculate if this is a multi-day range
    days_difference = (end_datetime - start_datetime).days
    
    # Define cheat_rate_params early so it's available for all cheat rate queries
    cheat_rate_params = [start_time, end_time]
    
    detailed_temporal_data = []
    if days_difference > 0:
        # Multi-day range: create detailed day+hour data (NON-CUMULATIVE - actual counts per hour)
        detailed_counts = {}
        
        # Initialize all day-hour combinations
        current_date = start_datetime.date()
        while current_date <= end_datetime.date():
            for hour in range(24):
                key = f"{current_date.strftime('%Y-%m-%d')}_{hour:02d}"
                detailed_counts[key] = 0
            current_date += timedelta(days=1)
        
        # Count actual analyses per time slot
        for analysis in all_analyses:
            try:
                analysis_datetime = datetime.strptime(analysis["date"], '%Y-%m-%d %H:%M:%S')
                analysis_date = analysis_datetime.date()
                analysis_hour = analysis_datetime.hour
                
                # Only include if within the filter range
                if start_datetime.date() <= analysis_date <= end_datetime.date():
                    key = f"{analysis_date.strftime('%Y-%m-%d')}_{analysis_hour:02d}"
                    if key in detailed_counts:
                        detailed_counts[key] += 1
            except Exception as e:
                logger.error(f"Error counting actual analyses: {e}")

        # Subtract verified analyses from their respective time slots
        for verification in verified_analyses:
            try:
                verification_datetime = datetime.strptime(verification["verification_timestamp"], '%Y-%m-%d %H:%M:%S')
                verification_date = verification_datetime.date()
                verification_hour = verification_datetime.hour
                
                # Only include if within the filter range
                if start_datetime.date() <= verification_date <= end_datetime.date():
                    key = f"{verification_date.strftime('%Y-%m-%d')}_{verification_hour:02d}"
                    if key in detailed_counts:
                        detailed_counts[key] = max(0, detailed_counts[key] - 1)
            except Exception as e:
                logger.error(f"Error counting verified analyses: {e}")

        # Format the detailed data (only include time slots that have data or are in business hours)
        for key, count in detailed_counts.items():
            date_str, hour_str = key.split('_')
            hour_int = int(hour_str)
            
            # Only include if there's data OR it's during business hours (8-18)
            if count > 0 or (8 <= hour_int <= 18):
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                day_name = english_to_french_day(date_obj.strftime('%A'))  # Day of week in French
                day_date = date_obj.strftime('%d/%m')  # DD/MM format
                
                detailed_temporal_data.append({
                    'date': date_str,
                    'day_name': day_name,
                    'day_date': day_date,
                    'hour': f"{hour_str}:00",
                    'count': count,
                        'label': f"{day_name} {day_date} {hour_str}:00",
                        'datetime_sort': f"{date_str} {hour_str}:00:00"
                })
        
        # Sort by actual datetime to ensure proper chronological order
        detailed_temporal_data.sort(key=lambda x: x['datetime_sort'])
    
    # 5.6. Get detailed cheat rate temporal distribution for multi-day ranges (NON-CUMULATIVE)
    detailed_cheat_rate_data = []
    if days_difference > 0:
        # Get detailed cheat rate data by day and hour
        detailed_cheat_rate_query = """
        SELECT 
            DATE(cr.date) as date,
            HOUR(cr.date) as hour,
            SUM(cr.nbr_etudiant) as total_students,
            SUM(cr.nbr_detection) as total_detections,
            CASE 
                WHEN SUM(cr.nbr_etudiant) > 0 
                THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
                ELSE 0 
            END as cheat_rate
        FROM cheat_rate cr"""
        
        detailed_cheat_rate_params = [start_time, end_time]
        
        # Add AREF filter if user is AREF admin
        if user_aref:
            detailed_cheat_rate_query += """
            JOIN lycee l ON cr.codeCentre = l.codeCentre
            JOIN ville v ON l.id_ville = v.id_ville
            JOIN dp d ON v.id_dp = d.id_dp
            JOIN aref a ON d.codeAref = a.codeAref"""
        
        detailed_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
        
        if user_aref:
            detailed_cheat_rate_query += " AND a.codeAref = %s"
            detailed_cheat_rate_params.append(user_aref)
        
        detailed_cheat_rate_query += """
        GROUP BY DATE(cr.date), HOUR(cr.date)
        ORDER BY date, hour
        """
        
        detailed_cheat_rate_result = db.execute_query(detailed_cheat_rate_query, tuple(detailed_cheat_rate_params))
        
        if detailed_cheat_rate_result is None:
            detailed_cheat_rate_result = []
        
        # Create a lookup for the detailed cheat rate data
        detailed_cheat_rate_lookup = {}
        for item in detailed_cheat_rate_result:
            key = f"{item['date']}_{item['hour']:02d}"
            detailed_cheat_rate_lookup[key] = item
        
        # Always include all day-hour slots in the filter range
        current_date = start_datetime.date()
        while current_date <= end_datetime.date():
            for hour in range(0, 24):
                key = f"{current_date.strftime('%Y-%m-%d')}_{hour:02d}"
                total_students = 0
                total_detections = 0
                cheat_rate = 0
                if key in detailed_cheat_rate_lookup:
                    item = detailed_cheat_rate_lookup[key]
                    total_students = item['total_students'] or 0
                    total_detections = item['total_detections'] or 0
                    cheat_rate = item['cheat_rate'] or 0
                date_obj = current_date
                day_name = english_to_french_day(date_obj.strftime('%A'))
                day_date = date_obj.strftime('%d/%m')
                detailed_cheat_rate_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day_name': day_name,
                    'day_date': day_date,
                    'hour': f"{hour:02d}:00",
                    'cheat_rate': round(cheat_rate, 2),
                    'total_students': total_students,
                    'total_detections': total_detections,
                    'label': f"{day_name} {day_date} {hour:02d}:00",
                    'datetime_sort': f"{current_date.strftime('%Y-%m-%d')} {hour:02d}:00:00"
                })
            current_date += timedelta(days=1)
        detailed_cheat_rate_data.sort(key=lambda x: x['datetime_sort'])

    # 13.5. Get hourly cheat rate distribution (national level) - CUMULATIVE
    hourly_cheat_rate_query = """
    SELECT 
        HOUR(cr.date) as hour,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr"""
    
    hourly_cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        hourly_cheat_rate_query += """
        JOIN lycee l ON cr.codeCentre = l.codeCentre
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref"""
    
    hourly_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
    
    if user_aref:
        hourly_cheat_rate_query += " AND a.codeAref = %s"
        hourly_cheat_rate_params.append(user_aref)
    
    hourly_cheat_rate_query += """
    GROUP BY HOUR(cr.date)
    ORDER BY hour
    """
    
    hourly_cheat_rate_data = db.execute_query(hourly_cheat_rate_query, tuple(hourly_cheat_rate_params))
    
    if hourly_cheat_rate_data is None:
        hourly_cheat_rate_data = []
    
    # Format hourly cheat rate data to include all hours in the filter range (CUMULATIVE)
    formatted_hourly_cheat_rate = []
    cumulative_students = 0
    cumulative_detections = 0
    latest_cheat_rate = 0  # Track the latest known cheat rate
    
    # Process every single hour in the time range, not just hours with data
    for hour in range(start_datetime.hour, end_datetime.hour + 1):
        # Check if we have data for this specific hour
        hour_data = next((item for item in hourly_cheat_rate_data if item['hour'] == hour), None)
        
        if hour_data:
            # We have data for this hour - add to cumulative totals
            cumulative_students += hour_data['total_students'] or 0
            cumulative_detections += hour_data['total_detections'] or 0
            latest_cheat_rate = (cumulative_detections / cumulative_students * 100) if cumulative_students > 0 else 0
        
        # Always add this hour to the result, using the latest known rate
        formatted_hourly_cheat_rate.append({
            'hour': f"{hour:02d}:00",
            'cheat_rate': round(latest_cheat_rate, 2),  # Use latest known rate for all hours
            'total_students': cumulative_students,
            'total_detections': cumulative_detections
        })
    
    # 6. Get day of week distribution
    day_of_week_counts = {
        "Monday": 0,
        "Tuesday": 0,
        "Wednesday": 0,
        "Thursday": 0,
        "Friday": 0,
        "Saturday": 0,
        "Sunday": 0
    }
    
    for analysis in all_analyses:
        try:
            day = datetime.strptime(analysis["date"], '%Y-%m-%d %H:%M:%S').strftime('%A')
            day_of_week_counts[day] += 1
        except Exception as e:
            logger.error(f"Error getting day of week distribution: {e}")
    formatted_day_of_week = []
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for day in days:
        formatted_day_of_week.append({
            'day': day,
            'count': day_of_week_counts[day]
        })
    
    # 7. Get counts by province (DP)
    province_counts = {}
    for analysis in all_analyses:
        dp = analysis["dp"]
        if dp not in province_counts:
            province_counts[dp] = 0
        province_counts[dp] += 1
    
    provinces = [{"name": dp, "count": count} for dp, count in province_counts.items()]
    provinces.sort(key=lambda x: x["count"], reverse=True)
    
    # 8. Get counts by exam center (lycee)
    center_counts = {}
    for analysis in all_analyses:
        lycee = analysis["lycee"]
        if lycee not in center_counts:
            center_counts[lycee] = 0
        center_counts[lycee] += 1
    
    centers = [{"name": lycee, "count": count} for lycee, count in center_counts.items()]
    centers.sort(key=lambda x: x["count"], reverse=True)
    
    # 9. Get counts by subject (matiere)
    subject_counts = {}
    for analysis in all_analyses:
        matiere = analysis["matiere"]
        if matiere not in subject_counts:
            subject_counts[matiere] = 0
        subject_counts[matiere] += 1
    
    subjects = [{"name": matiere, "count": count} for matiere, count in subject_counts.items()]
    subjects.sort(key=lambda x: x["count"], reverse=True)
    
    # 10. Get operateur distribution
    operateur_counts = {}
    for analysis in all_analyses:
        operateur = analysis["operateur"]
        if operateur not in operateur_counts:
            operateur_counts[operateur] = 0
        operateur_counts[operateur] += 1
    
    # 11. Get counts by verificateur
    verificateur_counts = {}
    for analysis in all_analyses:
        verificateur = analysis["verificateur_name"]
        if verificateur not in verificateur_counts:
            verificateur_counts[verificateur] = 0
        verificateur_counts[verificateur] += 1
    
    verificateurs = [{"name": verificateur, "count": count} for verificateur, count in verificateur_counts.items()]
    verificateurs.sort(key=lambda x: x["count"], reverse=True)
    
    # 12. Get cheat rate data by region (AREF)
    cheat_rate_query = """
    SELECT 
        a.name as aref_name,
        a.codeAref,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr
    JOIN lycee l ON cr.codeCentre = l.codeCentre
    JOIN ville v ON l.id_ville = v.id_ville
    JOIN dp d ON v.id_dp = d.id_dp
    JOIN aref a ON d.codeAref = a.codeAref
    WHERE cr.date BETWEEN %s AND %s"""
    
    cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        cheat_rate_query += " AND a.codeAref = %s"
        cheat_rate_params.append(user_aref)
    
    cheat_rate_query += """
    GROUP BY a.codeAref, a.name
    ORDER BY cheat_rate DESC
    """
    
    cheat_rate_data = db.execute_query(cheat_rate_query, tuple(cheat_rate_params))
    
    if cheat_rate_data is None:
        cheat_rate_data = []
    
    # 13. Get national cheat rate over time
    national_cheat_rate_query = """
    SELECT 
        DATE(cr.date) as date,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr"""
    
    national_cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        national_cheat_rate_query += """
        JOIN lycee l ON cr.codeCentre = l.codeCentre
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref"""
    
    national_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
    
    if user_aref:
        national_cheat_rate_query += " AND a.codeAref = %s"
        national_cheat_rate_params.append(user_aref)
    
    national_cheat_rate_query += """
    GROUP BY DATE(cr.date)
    ORDER BY date
    """
    
    national_cheat_rate_data = db.execute_query(national_cheat_rate_query, tuple(national_cheat_rate_params))
    
    if national_cheat_rate_data is None:
        national_cheat_rate_data = []
    
    # 14. Get cheat rate by province
    province_cheat_rate_query = """
    SELECT 
        d.name as province_name,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr
    JOIN lycee l ON cr.codeCentre = l.codeCentre
    JOIN ville v ON l.id_ville = v.id_ville
    JOIN dp d ON v.id_dp = d.id_dp"""
    
    province_cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        province_cheat_rate_query += """
        JOIN aref a ON d.codeAref = a.codeAref"""
    
    province_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
    
    if user_aref:
        province_cheat_rate_query += " AND a.codeAref = %s"
        province_cheat_rate_params.append(user_aref)
    
    province_cheat_rate_query += """
    GROUP BY d.name
    ORDER BY cheat_rate DESC
    """
    
    province_cheat_rate_data = db.execute_query(province_cheat_rate_query, tuple(province_cheat_rate_params))
    
    if province_cheat_rate_data is None:
        province_cheat_rate_data = []
    
    # 15. Get cheat rate by exam center
    center_cheat_rate_query = """
    SELECT 
        l.name_fr as center_name,
        l.name_ar as center_name_ar,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr
    JOIN lycee l ON cr.codeCentre = l.codeCentre"""
    
    center_cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        center_cheat_rate_query += """
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref"""
    
    center_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
    
    if user_aref:
        center_cheat_rate_query += " AND a.codeAref = %s"
        center_cheat_rate_params.append(user_aref)
    
    center_cheat_rate_query += """
    GROUP BY l.codeCentre, l.name_fr, l.name_ar
    ORDER BY cheat_rate DESC
    LIMIT 20
    """
    
    center_cheat_rate_data = db.execute_query(center_cheat_rate_query, tuple(center_cheat_rate_params))
    
    if center_cheat_rate_data is None:
        center_cheat_rate_data = []
    
    # 16. Get cheat rate by session (matiere)
    session_cheat_rate_query = """
    SELECT 
        cr.matiere as session_name,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr"""
    
    session_cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        session_cheat_rate_query += """
        JOIN lycee l ON cr.codeCentre = l.codeCentre
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref"""
    
    session_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
    
    if user_aref:
        session_cheat_rate_query += " AND a.codeAref = %s"
        session_cheat_rate_params.append(user_aref)
    
    session_cheat_rate_query += """
    GROUP BY cr.matiere
    ORDER BY cheat_rate DESC
    """
    
    session_cheat_rate_data = db.execute_query(session_cheat_rate_query, tuple(session_cheat_rate_params))
    
    if session_cheat_rate_data is None:
        session_cheat_rate_data = []
    
    # 17. Get cheat rate by examen type
    examen_cheat_rate_query = """
    SELECT 
        cr.examen as examen_name,
        SUM(cr.nbr_etudiant) as total_students,
        SUM(cr.nbr_detection) as total_detections,
        CASE 
            WHEN SUM(cr.nbr_etudiant) > 0 
            THEN ROUND((SUM(cr.nbr_detection) / SUM(cr.nbr_etudiant)) * 100, 2)
            ELSE 0 
        END as cheat_rate
    FROM cheat_rate cr"""
    
    examen_cheat_rate_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        examen_cheat_rate_query += """
        JOIN lycee l ON cr.codeCentre = l.codeCentre
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref"""
    
    examen_cheat_rate_query += " WHERE cr.date BETWEEN %s AND %s"
    
    if user_aref:
        examen_cheat_rate_query += " AND a.codeAref = %s"
        examen_cheat_rate_params.append(user_aref)
    
    examen_cheat_rate_query += """
    GROUP BY cr.examen
    ORDER BY cheat_rate DESC
    """
    
    examen_cheat_rate_data = db.execute_query(examen_cheat_rate_query, tuple(examen_cheat_rate_params))
    
    if examen_cheat_rate_data is None:
        examen_cheat_rate_data = []
    
    # 18. Get hourly cheat rate distribution by region (for the temporal analysis chart)
    hourly_cheat_rate_by_region = []
    
    if user_aref:
        # For AREF admins: only show their region's cumulative hourly cheat rate data
        hourly_cheat_rate_region_query = """
        SELECT 
            HOUR(cr.date) as hour,
            cr.date,
            cr.nbr_etudiant,
            cr.nbr_detection
        FROM cheat_rate cr
        JOIN lycee l ON cr.codeCentre = l.codeCentre
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref
        WHERE cr.date BETWEEN %s AND %s
        AND a.codeAref = %s
        ORDER BY cr.date
        """
        
        hourly_cheat_rate_region_params = [start_time, end_time, user_aref]
        hourly_cheat_rate_region_result = db.execute_query(hourly_cheat_rate_region_query, tuple(hourly_cheat_rate_region_params))
        
        if hourly_cheat_rate_region_result:
            # Get AREF name for display
            aref_name_query = "SELECT name FROM aref WHERE codeAref = %s"
            aref_name_result = db.execute_query(aref_name_query, (user_aref,))
            aref_name = aref_name_result[0]['name'] if aref_name_result else f"AREF {user_aref}"
            
            # Calculate cumulative cheat rates up to each hour
            cumulative_students = {}
            cumulative_detections = {}
            
            # Initialize for all hours in the filter range
            for hour in range(start_datetime.hour, end_datetime.hour + 1):
                cumulative_students[hour] = 0
                cumulative_detections[hour] = 0
            
            # Process data chronologically and accumulate
            for item in hourly_cheat_rate_region_result:
                hour = item['hour']
                students = item['nbr_etudiant'] or 0
                detections = item['nbr_detection'] or 0
                
                # Add to all hours from this hour onwards (cumulative effect)
                for h in range(hour, end_datetime.hour + 1):
                    if h in cumulative_students:
                        cumulative_students[h] += students
                        cumulative_detections[h] += detections
            
            # Format data for the region with cumulative rates
            region_hourly_data = []
            for hour in range(start_datetime.hour, end_datetime.hour + 1):
                total_students = cumulative_students[hour]
                total_detections = cumulative_detections[hour]
                
                if total_students > 0:
                    cheat_rate = round((total_detections / total_students) * 100, 2)
                else:
                    cheat_rate = 0
                
                region_hourly_data.append({
                    'hour': f"{hour:02d}:00",
                    'cheat_rate': cheat_rate
                })
            
            hourly_cheat_rate_by_region.append({
                'region': aref_name,
                'data': region_hourly_data
            })
    else:
        # For super admins: show all regions' cumulative hourly cheat rate data + global line
        hourly_cheat_rate_all_regions_query = """
        SELECT 
            a.name as aref_name,
            a.codeAref as aref_code,
            HOUR(cr.date) as hour,
            cr.date,
            cr.nbr_etudiant,
            cr.nbr_detection
        FROM cheat_rate cr
        JOIN lycee l ON cr.codeCentre = l.codeCentre
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref
        WHERE cr.date BETWEEN %s AND %s
        ORDER BY a.name, cr.date
        """
        
        hourly_cheat_rate_all_regions_params = [start_time, end_time]
        hourly_cheat_rate_all_regions_result = db.execute_query(hourly_cheat_rate_all_regions_query, tuple(hourly_cheat_rate_all_regions_params))
        
        if hourly_cheat_rate_all_regions_result:
            # Group by region
            regions_data = {}
            for item in hourly_cheat_rate_all_regions_result:
                aref_name = item['aref_name']
                if aref_name not in regions_data:
                    regions_data[aref_name] = []
                regions_data[aref_name].append(item)
            
            # Calculate cumulative data for each region
            for aref_name, region_results in regions_data.items():
                # Calculate cumulative cheat rates up to each hour for this region
                cumulative_students = {}
                cumulative_detections = {}
                
                # Initialize for all hours in the filter range
                for hour in range(start_datetime.hour, end_datetime.hour + 1):
                    cumulative_students[hour] = 0
                    cumulative_detections[hour] = 0
                
                # Process data chronologically and accumulate
                for item in region_results:
                    hour = item['hour']
                    students = item['nbr_etudiant'] or 0
                    detections = item['nbr_detection'] or 0
                    
                    # Add to all hours from this hour onwards (cumulative effect)
                    for h in range(hour, end_datetime.hour + 1):
                        if h in cumulative_students:
                            cumulative_students[h] += students
                            cumulative_detections[h] += detections
                
                # Format data for the region with cumulative rates
                region_hourly_data = []
                for hour in range(start_datetime.hour, end_datetime.hour + 1):
                    total_students = cumulative_students[hour]
                    total_detections = cumulative_detections[hour]
                    
                    if total_students > 0:
                        cheat_rate = round((total_detections / total_students) * 100, 2)
                    else:
                        cheat_rate = 0
                    
                    region_hourly_data.append({
                        'hour': f"{hour:02d}:00",
                        'cheat_rate': cheat_rate
                    })
                
                hourly_cheat_rate_by_region.append({
                    'region': aref_name,
                    'data': region_hourly_data
                })
            
            # Add global line (cumulative average of all regions)
            global_cumulative_students = {}
            global_cumulative_detections = {}
            
            # Initialize for all hours in the filter range
            for hour in range(start_datetime.hour, end_datetime.hour + 1):
                global_cumulative_students[hour] = 0
                global_cumulative_detections[hour] = 0
            
            # Accumulate data from all regions
            for item in hourly_cheat_rate_all_regions_result:
                hour = item['hour']
                students = item['nbr_etudiant'] or 0
                detections = item['nbr_detection'] or 0
                
                # Add to all hours from this hour onwards (cumulative effect)
                for h in range(hour, end_datetime.hour + 1):
                    if h in global_cumulative_students:
                        global_cumulative_students[h] += students
                        global_cumulative_detections[h] += detections
            
            # Calculate global cumulative rates
            global_hourly_data = []
            for hour in range(start_datetime.hour, end_datetime.hour + 1):
                total_students = global_cumulative_students[hour]
                total_detections = global_cumulative_detections[hour]
                
                if total_students > 0:
                    cheat_rate = round((total_detections / total_students) * 100, 2)
                else:
                    cheat_rate = 0
                
                global_hourly_data.append({
                    'hour': f"{hour:02d}:00",
                    'cheat_rate': cheat_rate
                })
            
            # Add global line to the beginning of the list
            hourly_cheat_rate_by_region.insert(0, {
                'region': 'National (Moyenne)',
                'data': global_hourly_data
            })

    # Calculate summary data for risk levels (this is for the summary cards)
    summary_data = {
        "yellow": total_general,  # All general analyses are considered "yellow" (potential risk)
        "orange": 0,  # This would be from mobility analyses that are pending verification
        "red": verified_count,     # This would be from verified mobility analyses (confirmed)
        "gray": denied_count,    # This would be from verified analyses marked as false alarms (denied)
        "total": total_general + total_mobility + total_verified
    }
    
    # For now, we'll count mobility analyses as "orange" (pending verification)
    summary_data["orange"] = total_mobility
    
    # Build communication types data structure
    communication_types_data = {
        "vocal": {
                "yellow": vocal_count,
                "orange": 0,
                "red": 0
            },
        "data": {
                "yellow": data_count,
                "orange": 0,
                "red": verified_count  # Verified analyses are typically data-based
        },
        "whatsapp": {
            "yellow": 0,
                "orange": 0,
                "red": 0
            }
    }
    
    return {
        "summaryData": summary_data,
        "regionData": regions,
        "timeOfDayData": time_of_day,
        "communicationTypesData": communication_types_data,
        "hourlyDistributionData": formatted_hourly,
        "dayOfWeekData": formatted_day_of_week,
        "provincesData": provinces,
        "examCentersData": centers,
        "subjectWarningsData": subjects,
        "communicationDistributionData": {
            "vocal": vocal_count,
            "data": data_count,
            "whatsapp": 0
        },
        "operateurDistributionData": operateur_counts,
        "verifiersData": verificateurs,
        "cheatRateData": {
            "byRegion": cheat_rate_data,
            "nationalOverTime": national_cheat_rate_data,
            "byProvince": province_cheat_rate_data,
            "byCenter": center_cheat_rate_data,
            "hourlyDistribution": formatted_hourly_cheat_rate,
            "session": session_cheat_rate_data,
            "examen": examen_cheat_rate_data
        },
        "detailedTemporalData": detailed_temporal_data,
        "detailedCheatRateData": detailed_cheat_rate_data,
        "hourlyDistributionByRegion": hourly_distribution_by_region,
        "hourlyCheatRateByRegion": hourly_cheat_rate_by_region
    }


@app.get("/api/export-data")
def export_data(start_time: Optional[str] = None, end_time: Optional[str] = None, request: Request = None):
    """Export analyses data and cheat rates as Excel file with two sheets"""
    
    # Default to today 8am-6pm if no time range provided (GMT+1)
    if not start_time and not end_time:
        today = get_morocco_date_str()
        start_time = f"{today} 08:00:00"
        end_time = f"{today} 18:00:00"
    
    # Get user AREF filter
    user_aref = get_user_aref_filter(request) if request else None
    
    # Log the received parameters for debugging

    # 1. Get general analyses (all data, not just latest batch) with AREF filtering
    general_query = """
    SELECT ag.id_analyse, ag.date, ag.operateur, ag.type_communication, ag.salle, ag.matiere, ag.batch,
           ag.aref, ag.dp, ag.ville, ag.lycee, v.full_name as verificateur_name
    FROM analyse_generale ag
    JOIN verificateur v ON ag.id_verificateur = v.id_verificateur
    WHERE ag.date BETWEEN %s AND %s
    """
    general_params = [start_time, end_time]
    if user_aref:
        general_query += " AND ag.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        general_params.append(user_aref)
    general_query += " ORDER BY ag.date DESC"
    general_analyses = db.execute_query(general_query, tuple(general_params)) or []

    # Batch filtering for general analyses: keep only latest batch per location per day
    filtered_general_analyses = []
    latest_batch_map = {}
    for analysis in general_analyses:
        date_str = str(analysis['date']).split(' ')[0]
        key = (analysis['aref'], analysis['dp'], analysis['ville'], analysis['lycee'], analysis['salle'], analysis['matiere'], date_str)
        batch = analysis.get('batch', 1) or 1
        if key not in latest_batch_map or batch > latest_batch_map[key][0]:
            latest_batch_map[key] = (batch, analysis)
    filtered_general_analyses = [v[1] for v in latest_batch_map.values()]
    
    # 2. Get mobility analyses (all data, not just latest batch) with AREF filtering
    mobility_query = """
    SELECT amf.id_analyse, amf.timestamp as date, amf.operator as operateur, amf.cne, amf.salle, amf.matiere, amf.batch,
           amf.aref, amf.dp, amf.ville, amf.lycee, v.full_name as verificateur_name
    FROM analyse_mobility_final amf
    JOIN verificateur v ON amf.id_verificateur = v.id_verificateur
    WHERE amf.timestamp BETWEEN %s AND %s
    """
    mobility_params = [start_time, end_time]
    if user_aref:
        mobility_query += " AND amf.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        mobility_params.append(user_aref)
    mobility_query += " ORDER BY amf.timestamp DESC"
    mobility_analyses = db.execute_query(mobility_query, tuple(mobility_params)) or []

    # Batch filtering for mobility analyses: keep only latest batch per location+CNE per day
    filtered_mobility_analyses = []
    latest_mobility_batch_map = {}
    for analysis in mobility_analyses:
        date_str = str(analysis['date']).split(' ')[0]
        key = (analysis['aref'], analysis['dp'], analysis['ville'], analysis['lycee'], analysis['salle'], analysis['matiere'], analysis['cne'], date_str)
        batch = analysis.get('batch', 1) or 1
        if key not in latest_mobility_batch_map or batch > latest_mobility_batch_map[key][0]:
            latest_mobility_batch_map[key] = (batch, analysis)
    filtered_mobility_analyses = [v[1] for v in latest_mobility_batch_map.values()]
    
    # 3. Get verified analyses (with complete info from mobility table) with AREF filtering
    verified_query = """
    SELECT av.id_analyse_verifier, av.action, av.timestamp as verification_date,
           amf.id_analyse, amf.timestamp as date, amf.operator as operateur, amf.cne, amf.salle, amf.matiere, amf.batch,
           amf.aref, amf.dp, amf.ville, amf.lycee, v.full_name as verificateur_name
    FROM analyse_verifier av
    JOIN analyse_mobility_final amf ON av.id_analyse_mobilite_final = amf.id_analyse
    JOIN verificateur v ON amf.id_verificateur = v.id_verificateur
    WHERE amf.timestamp BETWEEN %s AND %s
    """
    verified_params = [start_time, end_time]
    if user_aref:
        verified_query += " AND amf.aref COLLATE utf8mb4_0900_ai_ci = (SELECT name FROM aref WHERE codeAref = %s)"
        verified_params.append(user_aref)
    verified_query += " ORDER BY av.timestamp DESC"
    verified_analyses = db.execute_query(verified_query, tuple(verified_params)) or []

    # Batch filtering for verified analyses: keep only latest verification per id_analyse_mobilite_final
    latest_verified_map = {}
    for analysis in verified_analyses:
        key = analysis['id_analyse']
        verif_id = analysis['id_analyse_verifier']
        if key not in latest_verified_map or verif_id > latest_verified_map[key]['id_analyse_verifier']:
            latest_verified_map[key] = analysis
    filtered_verified_analyses = list(latest_verified_map.values())

    # Use filtered lists in the export logic below (replace general_analyses, mobility_analyses, verified_analyses)
    general_analyses = filtered_general_analyses
    mobility_analyses = filtered_mobility_analyses
    verified_analyses = filtered_verified_analyses
    
    # 4. Get cheat rates data with AREF filtering
    cheat_rates_query = """
    SELECT cr.*, l.name_fr as lycee_name, l.name_ar as lycee_name_ar,
           v.name as ville_name, dp.name as dp_name, aref.name as aref_name
    FROM cheat_rate cr
    JOIN lycee l ON cr.codeCentre = l.codeCentre
    LEFT JOIN ville v ON l.id_ville = v.id_ville
    LEFT JOIN dp ON v.id_dp = dp.id_dp
    LEFT JOIN aref ON dp.codeAref = aref.codeAref
    WHERE cr.date BETWEEN %s AND %s
    """
    cheat_rates_params = [start_time, end_time]
    
    # Add AREF filter if user is AREF admin
    if user_aref:
        cheat_rates_query += " AND aref.codeAref = %s"
        cheat_rates_params.append(user_aref)
    
    cheat_rates_query += " ORDER BY cr.date DESC"
    
    cheat_rates_data = db.execute_query(cheat_rates_query, tuple(cheat_rates_params))
    if cheat_rates_data is None:
        cheat_rates_data = []
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove the default and cheat rate sheets if they exist
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in ("Analyses",):
            wb.remove(wb[sheet_name])
    if 'Analyses' in wb.sheetnames:
        wb.remove(wb['Analyses'])
    
    # Create single sheet for analyses data
    ws1 = wb.create_sheet("Analyses")
    
    # Headers for analyses sheet (with 'N° d\'examen' instead of 'CNE', and no ID, no verification columns)
    headers = [
        "Type de Risque", "Date", "AREF", "Province", "Ville", "Centre d'Examen", "Salle", "Session", 
        "Vérificateur", "N° d'examen", "Nbr Détections"
    ]
    
    # Add headers to first sheet
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row_num = 2
    # Add mobility analyses data (Annoté)
    for analysis in mobility_analyses:
        ws1.cell(row=row_num, column=1, value="Risque Annoté")
        ws1.cell(row=row_num, column=2, value=str(analysis["date"]))
        ws1.cell(row=row_num, column=3, value=analysis["aref"] or "")
        ws1.cell(row=row_num, column=4, value=analysis["dp"] or "")
        ws1.cell(row=row_num, column=5, value=analysis["ville"] or "")
        ws1.cell(row=row_num, column=6, value=analysis["lycee"] or "")
        ws1.cell(row=row_num, column=7, value=analysis["salle"] or "")
        ws1.cell(row=row_num, column=8, value=analysis["matiere"] or "")
        ws1.cell(row=row_num, column=9, value=analysis["verificateur_name"] or "")
        ws1.cell(row=row_num, column=10, value=analysis["cne"] or "")  # N° d'examen
        ws1.cell(row=row_num, column=11, value=analysis["batch"] or "")
        row_num += 1
    
    # Add verified/denied analyses data (Vérifié, Fausse Alerte)
    for analysis in verified_analyses:
        risk_type = "Risque Confirmé" if analysis["action"] == "confirm" else "Fausse Alerte"
        ws1.cell(row=row_num, column=1, value=risk_type)
        ws1.cell(row=row_num, column=2, value=str(analysis["date"]))
        ws1.cell(row=row_num, column=3, value=analysis["aref"] or "")
        ws1.cell(row=row_num, column=4, value=analysis["dp"] or "")
        ws1.cell(row=row_num, column=5, value=analysis["ville"] or "")
        ws1.cell(row=row_num, column=6, value=analysis["lycee"] or "")
        ws1.cell(row=row_num, column=7, value=analysis["salle"] or "")
        ws1.cell(row=row_num, column=8, value=analysis["matiere"] or "")
        ws1.cell(row=row_num, column=9, value=analysis["verificateur_name"] or "")
        ws1.cell(row=row_num, column=10, value=analysis["cne"] or "")  # N° d'examen
        ws1.cell(row=row_num, column=11, value=analysis["batch"] or "")
        row_num += 1
    
    # Auto-adjust column widths for the sheet
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate custom filename based on user type and date range
    now = get_morocco_time()
    # Use timestamp format with current date (e.g., "26-06-14h30")
    current_date = now.strftime('%d-%m')
    current_time = now.strftime('%Hh%M')
    timestamp_export = f"{current_date}-{current_time}"
    
    # Format start and end dates for filename - use day-month format without year
    try:
        start_date_formatted = datetime.strptime(start_time.split(' ')[0], '%Y-%m-%d').strftime('%d-%m')
        end_date_formatted = datetime.strptime(end_time.split(' ')[0], '%Y-%m-%d').strftime('%d-%m')
    except:
        # Fallback if date parsing fails - use current date in day-month format
        fallback_date = now.strftime('%d-%m')
        start_date_formatted = start_time.split(' ')[0] if start_time else fallback_date
        end_date_formatted = end_time.split(' ')[0] if end_time else fallback_date
        # If we have full dates, try to extract day-month
        if start_date_formatted and len(start_date_formatted) > 5:
            try:
                start_date_formatted = datetime.strptime(start_date_formatted, '%Y-%m-%d').strftime('%d-%m')
            except:
                start_date_formatted = fallback_date
        if end_date_formatted and len(end_date_formatted) > 5:
            try:
                end_date_formatted = datetime.strptime(end_date_formatted, '%Y-%m-%d').strftime('%d-%m')
            except:
                end_date_formatted = fallback_date
    
    if user_aref:
        # AREF admin - get AREF name and abbreviation
        aref_query = "SELECT name, abbreviation FROM aref WHERE codeAref = %s"
        aref_result = db.execute_query(aref_query, (user_aref,))
        
        if aref_result:
            aref_name = aref_result[0]['name']
            aref_abbreviation = aref_result[0]['abbreviation']
            # Use AREF abbreviation for filename
            filename = f"T3S-{aref_abbreviation.lower()}-de-{start_date_formatted}-à-{end_date_formatted}-généré_le_{timestamp_export}.xlsx"
        else:
            # Fallback if AREF name not found
            filename = f"t3-shield-rapport-aref-{user_aref}-de-{start_date_formatted}-à-{end_date_formatted}-généré_le_{timestamp_export}.xlsx"
    else:
        # Super admin - national report
        filename = f"t3-shield-rapport-national-de-{start_date_formatted}-à-{end_date_formatted}-généré_le_{timestamp_export}.xlsx"
    
    # Return Excel file
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )
        
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        # Send a welcome message
        await websocket.send_text(json.dumps({"event": "connected", "message": "Connected to WebSocket server"}))
        
        # Keep the connection alive with a ping-pong mechanism
        while True:
            try:
                # Wait for messages with a timeout
                data = await asyncio.wait_for(websocket.receive_text(), timeout=30)
                # Echo back any received messages
                await websocket.send_text(data)
            except asyncio.TimeoutError:
                # Send a ping message to keep the connection alive
                await websocket.send_text(json.dumps({"event": "ping"}))
    except WebSocketDisconnect:
        manager.disconnect(websocket)


app.mount("/static", StaticFiles(directory="front-end", html=True), name="static")
app.mount("/css", StaticFiles(directory="front-end/css"), name="css")
app.mount("/js", StaticFiles(directory="front-end/js"), name="js")

# Serve the dashboard at the root of /dashboard-national
@app.get("/dashboard-national", include_in_schema=False)
def get_dashboard():
    return FileResponse("front-end/index.html")
# Add this endpoint to serve the stats dashboard
@app.get("/stats-dashboard", include_in_schema=False)
def get_stats_dashboard():
    return FileResponse("front-end/stats.html")

@app.get("/statistiques", include_in_schema=False)
def get_stats_refined():
    return FileResponse("front-end/stats-refined.html")

@app.get("/t3shield/api/get_arefs")
def get_arefs(request: Request):
    try:
        db = Database()
        db.connect()
        
        # Get AREF filter from session (returns None if not authenticated)
        aref_filter = get_user_aref_filter(request)
        
        if aref_filter:
            # AREF admin - only return their AREF
            query = """
            SELECT a.id_aref as id_aref, a.name as name, a.coordinates as coordinates, a.codeAref as codeAref, a.abbreviation as abbreviation
            FROM aref a
            WHERE a.codeAref = %s
            """
            results = db.execute_query(query, (aref_filter,))
        else:
            # Super admin or unauthenticated - return all AREFs
            query = """
            SELECT a.id_aref as id_aref, a.name as name, a.coordinates as coordinates, a.codeAref as codeAref, a.abbreviation as abbreviation
            FROM aref a
            """
            results = db.execute_query(query)
        
        arefs = []
        for row in results:
            # Convert coordinates string to array format
            coords = row["coordinates"].split(',')
            coords = [float(coord.strip()) for coord in coords]
            
            aref = {
                "id": f"aref-{row['id_aref']}",  # Format: aref-arefId
                "name": row["name"],
                "type": "AREF",
                "coordinates": coords,
                "additionalInfo": {
                    "description": f"Académie Régionale d'Education et de Formation {row['name']}",
                    "codeAref": row["codeAref"],
                    "abbreviation": row["abbreviation"]
                }
            }
            arefs.append(aref)
        
        if arefs:
            logger.info(f"Found {len(arefs)} AREFs")
        return arefs
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/t3shield/api/get_dps")
def get_dps(request: Request):
    try:
        db = Database()
        db.connect()
        
        # Get AREF filter from session (returns None if not authenticated)
        aref_filter = get_user_aref_filter(request)
        
        if aref_filter:
            # AREF admin - only return DPs for their AREF
            query = """
            SELECT d.id_dp as id_dp, d.name as name, d.coordinates as coordinates, d.codeAref as codeAref, a.id_aref as id_aref
            FROM dp d
            JOIN aref a ON d.codeAref = a.codeAref
            WHERE a.codeAref = %s
            """
            results = db.execute_query(query, (aref_filter,))
        else:
            # Super admin or unauthenticated - return all DPs
            query = """
            SELECT d.id_dp as id_dp, d.name as name, d.coordinates as coordinates, d.codeAref as codeAref, a.id_aref as id_aref
            FROM dp d
            JOIN aref a ON d.codeAref = a.codeAref
            """
            results = db.execute_query(query)
        
        dps = []
        for row in results:
            # Convert coordinates string to array format
            coords = row["coordinates"].split(',')
            coords = [float(coord.strip()) for coord in coords]
            
            dp = {
                "id": f"dp-{row['id_dp']}-{row['id_aref']}",  # Format: dp-arefId-dpId
                "name": row["name"],
                "type": "DP",
                "parentId": f"aref-{row['id_aref']}",  # Link to parent AREF
                "coordinates": coords,
                "additionalInfo": {
                    "description": f"Direction Provinciale de {row['name']}"
                }
            }
            dps.append(dp)
        
        if dps:
            logger.info(f"Found {len(dps)} DPs")
        return dps
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
    finally:
        db.close()

@app.get("/t3shield/api/get_villes")
def get_villes(request: Request):
    try:
        db = Database()
        db.connect()
        
        # Get AREF filter from session (returns None if not authenticated)
        aref_filter = get_user_aref_filter(request)
        
        if aref_filter:
            # AREF admin - only return villes for their AREF
            query = """
            SELECT v.id_ville as id_ville, v.name as name, v.coordinates as coordinates, v.id_dp as id_dp, d.id_dp as dp_id, a.id_aref as id_aref
            FROM ville v
            JOIN dp d ON v.id_dp = d.id_dp
            JOIN aref a ON d.codeAref = a.codeAref
            WHERE a.codeAref = %s
            """
            results = db.execute_query(query, (aref_filter,))
        else:
            # Super admin or unauthenticated - return all villes
            query = """
            SELECT v.id_ville as id_ville, v.name as name, v.coordinates as coordinates, v.id_dp as id_dp, d.id_dp as dp_id, a.id_aref as id_aref
            FROM ville v
            JOIN dp d ON v.id_dp = d.id_dp
            JOIN aref a ON d.codeAref = a.codeAref
            """
            results = db.execute_query(query)
        
        villes = []
        for row in results:
            # Convert coordinates string to array format
            coords = row["coordinates"].split(',')
            coords = [float(coord.strip()) for coord in coords]
            
            ville = {
                "id": f"ville-{row['id_ville']}-{row['id_dp']}-{row['id_aref']}",  # Format: ville-arefId-dpId-villeId
                "name": row["name"],
                "type": "VILLE",
                "parentId": f"dp-{row['id_dp']}-{row['id_aref']}",  # Link to parent DP
                "coordinates": coords,
                "additionalInfo": {
                    "description": f"Ville de {row['name']}"
                }
            }
            villes.append(ville)
        
        if villes:
            logger.info(f"Found {len(villes)} Villes")
        return villes
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/t3shield/api/get_lycees")
def get_lycees(request: Request):
    try:
        db = Database()
        db.connect()
        
        # Get AREF filter from session (returns None if not authenticated)
        aref_filter = get_user_aref_filter(request)
        
        if aref_filter:
            # AREF admin - only return lycees for their AREF
            query = """
            SELECT l.id_lycee as id_lycee, l.name_fr as name, l.coordinates as coordinates, l.id_ville as id_ville, v.id_ville as ville_id, 
                d.id_dp as id_dp, a.id_aref as id_aref
            FROM lycee l
            JOIN ville v ON l.id_ville = v.id_ville
            JOIN dp d ON v.id_dp = d.id_dp
            JOIN aref a ON d.codeAref = a.codeAref
            WHERE a.codeAref = %s
            """
            results = db.execute_query(query, (aref_filter,))
        else:
            # Super admin or unauthenticated - return all lycees
            query = """
            SELECT l.id_lycee as id_lycee, l.name_fr as name, l.coordinates as coordinates, l.id_ville as id_ville, v.id_ville as ville_id, 
                   d.id_dp as id_dp, a.id_aref as id_aref
            FROM lycee l
            JOIN ville v ON l.id_ville = v.id_ville
            JOIN dp d ON v.id_dp = d.id_dp
            JOIN aref a ON d.codeAref = a.codeAref
            """
            results = db.execute_query(query)
        
        lycees = []
        for row in results:
            # Convert coordinates string to array format
            coords = row["coordinates"].split(',')
            coords = [float(coord.strip()) for coord in coords]
            
            lycee = {
                "id": f"lycee-{row['id_lycee']}-{row['id_ville']}-{row['id_dp']}-{row['id_aref']}",  # Format: lycee-arefId-dpId-villeId-lyceeId
                "name": row["name"],
                "type": "LYCEE",
                "parentId": f"ville-{row['id_ville']}-{row['id_dp']}-{row['id_aref']}",  # Link to parent Ville
                "coordinates": coords,
                "additionalInfo": {
                    "description": f"Lycée {row['name']}",
                    "address": row["name"]  # Using name as address since we don't have address in DB
                }
            }
            
            lycees.append(lycee)
        
        if lycees:
            logger.info(f"Found {len(lycees)} Lycees")
        return lycees
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

# Chef Center endpoints
@app.get("/chef/centre/{code_centre}")
def chef_center_access(code_centre: str, request: Request):
    """Check if center is active and serve appropriate page"""
    try:
        # Check if center is active
        query = "SELECT active, username, url FROM chef_center WHERE codeCentre = %s"
        result = db.execute_query(query, (code_centre,))
        
        if not result:
            # Center not found - serve no radar page
            return FileResponse("front-end/no-radar.html")
        
        center = result[0]
        if not center["active"]:
            # Center inactive - serve no radar page
            return FileResponse("front-end/no-radar.html")
        
        # Center is active - check if user is already logged in
        chef_session_token = request.cookies.get("chef_centre_session")
        if chef_session_token and chef_session_token in chef_centre_sessions:
            session_data = chef_centre_sessions[chef_session_token]
            if session_data["code_centre"] == code_centre:
                # User is already logged in, redirect to dashboard
                return RedirectResponse(url=f"/chef/centre/{code_centre}/dashboard")
        
        # User not logged in - serve login page
        return FileResponse("front-end/chef-login.html")
        
    except Exception as e:
        logger.error(f"Error in chef_center_access: {str(e)}")
        return FileResponse("front-end/no-radar.html")

@app.post("/chef/centre/{code_centre}/login")
def chef_center_login(code_centre: str, login_data: dict, response: Response):
    """Authenticate chef center"""
    try:
        username = login_data.get("username")
        password = login_data.get("password")
        
        if not username or not password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Username and password required"
            )
        
        # Hash password and check credentials
        hashed_password = password #hash_password(password)
        query = """
        SELECT id_chef_centre, username, url, id_verificateur 
        FROM chef_center 
        WHERE codeCentre = %s AND username = %s AND password = %s AND active = 1
        """
        
        result = db.execute_query(query, (code_centre, username, hashed_password))

        if not result:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Identifiants incorrects ou centre inactif"
            )
        
        center = result[0]
        
        # Clean up expired sessions first
        cleanup_expired_sessions()
        
        # Clear any existing sessions for this centre to prevent conflicts
        sessions_to_remove = []
        for token, session_data in chef_centre_sessions.items():
            if session_data["code_centre"] == code_centre:
                sessions_to_remove.append(token)
        
        for token in sessions_to_remove:

            del chef_centre_sessions[token]
        
        # Create new session
        session_token = str(uuid.uuid4())
        chef_centre_sessions[session_token] = {
            "code_centre": code_centre,
            "center_data": center,
            "login_time": get_morocco_time()
        }
        
        # Set session cookie
        response.set_cookie(
            key="chef_centre_session",
            value=session_token,
            httponly=True,
            secure=False,  # Set to True in production with HTTPS
            samesite="lax",
            max_age=3600 * 8,  # 8 hours
            path="/"  # Ensure cookie is available for all paths
        )
        
        return {
            "status": "success",
            "message": "Authentification réussie",
            "center": center
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in chef_center_login: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur serveur"
        )

@app.post("/chef/centre/{code_centre}/logout")
def chef_center_logout(code_centre: str, request: Request, response: Response):
    """Logout chef center - Complete session cleanup"""

    # Get session token from cookies (try all possible cookie names)
    session_token = (
        request.cookies.get("chef_centre_session") or 
        request.cookies.get(f"chef_session_{code_centre}") or
        request.cookies.get("session")
    )


    # Remove the specific session if it exists
    if session_token and session_token in chef_centre_sessions:

        del chef_centre_sessions[session_token]
    
    # Also remove any sessions for this centre code as a safety measure
    sessions_to_remove = []
    for token, session_data in chef_centre_sessions.items():
        if session_data.get("code_centre") == code_centre:
            sessions_to_remove.append(token)
    
    for token in sessions_to_remove:

        del chef_centre_sessions[token]
    
    # Clear ALL possible cookie variations
    cookie_variations = [
        "chef_centre_session",
        f"chef_session_{code_centre}",
        "session",
        "chef_session"
    ]
    
    path_variations = [
        "/",
        f"/chef/centre/{code_centre}",
        f"/chef/centre/{code_centre}/",
        f"/chef"
    ]
    
    # Clear cookies with all path variations
    for cookie_name in cookie_variations:
        for path in path_variations:
            response.delete_cookie(
                key=cookie_name,
                path=path,
                domain=None,
                secure=False,
                httponly=True,
                samesite="lax"
            )


    return {"status": "success", "message": "Déconnexion réussie"}

@app.get("/chef/centre/{code_centre}/dashboard")
def chef_center_dashboard(code_centre: str):
    """Serve chef center dashboard"""
    return FileResponse("front-end/chef-dashboard.html")

@app.get("/chef/centre/{code_centre}/api/info")
def get_chef_center_info(code_centre: str):
    """Get center information including name_fr"""
    try:
        # Get center info with verificateur details
        center_query = """
        SELECT cc.codeCentre, cc.username, v.full_name, v.aref, v.dp, v.ville, v.lycee
        FROM chef_center cc
        JOIN verificateur v ON cc.id_verificateur = v.id_verificateur
        WHERE cc.codeCentre = %s AND cc.active = 1
        """
        center_result = db.execute_query(center_query, (code_centre,))
        
        if not center_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Centre non trouvé ou inactif"
            )
        
        center_info = center_result[0]
        return {
            "code_centre": center_info["codeCentre"],
            "username": center_info["username"],
            "name_fr": center_info["lycee"],  # Using lycee as the center name
            "full_name": center_info["full_name"],
            "aref": center_info["aref"],
            "dp": center_info["dp"],
            "ville": center_info["ville"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_chef_center_info: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la récupération des informations du centre"
        )

@app.get("/chef/centre/{code_centre}/api/analyses")
def get_chef_center_analyses(code_centre: str):
    """Get analyses for specific center with batch filtering"""
    try:
        # Get center's location data from codeCentre
        lycee_query = """
        SELECT l.name_fr, l.id_ville, v.name as ville_name, d.name as dp_name, a.name as aref_name
        FROM lycee l
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref
        WHERE l.codeCentre = %s
        """
        lycee_result = db.execute_query(lycee_query, (code_centre,))
        
        if not lycee_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Centre non trouvé"
            )
        
        center_data = lycee_result[0]
        lycee_name = center_data["name_fr"]
        aref_name = center_data["aref_name"]
        dp_name = center_data["dp_name"]
        ville_name = center_data["ville_name"]
        
        # Get general analyses for this center (filtered by center location)
        general_query = """
        SELECT ag.id_analyse, ag.date, ag.operateur, ag.type_communication, ag.salle, ag.matiere, ag.batch
        FROM analyse_generale ag
        WHERE ag.aref = %s AND ag.dp = %s AND ag.ville = %s AND ag.lycee = %s
        ORDER BY ag.date DESC
        """
        general_analyses_raw = db.execute_query(general_query, (aref_name, dp_name, ville_name, lycee_name))
        
        # Get mobility analyses for this center (filtered by center location)
        mobility_query = """
        SELECT amf.id_analyse, amf.timestamp, amf.cne, amf.salle, amf.matiere, amf.batch,
               av.action as verification_action
        FROM analyse_mobility_final amf
        LEFT JOIN analyse_verifier av ON amf.id_analyse = av.id_analyse_mobilite_final
        WHERE amf.aref = %s AND amf.dp = %s AND amf.ville = %s AND amf.lycee = %s
        ORDER BY amf.timestamp DESC
        """
        mobility_analyses_raw = db.execute_query(mobility_query, (aref_name, dp_name, ville_name, lycee_name))
        
        # Apply date-aware batch filtering for general analyses
        location_batches = {}
        filtered_general_analyses = []
        
        if general_analyses_raw:
            # First pass: find the highest batch for each location per day
            for analysis in general_analyses_raw:
                # Extract date from date field for date-aware grouping
                date_field = analysis['date']
                if isinstance(date_field, str):
                    analysis_date = date_field.split(' ')[0]  # Get YYYY-MM-DD part
                else:
                    analysis_date = date_field.strftime('%Y-%m-%d')
                
                location_key = f"{analysis['salle']}-{analysis['matiere']}-{analysis_date}"
                batch = analysis.get('batch', 1)
                if batch is None:
                    batch = 1
                
                if location_key not in location_batches or batch > location_batches[location_key]:
                    location_batches[location_key] = batch
            
            # Second pass: keep only analyses with the highest batch for each location per day
            for analysis in general_analyses_raw:
                # Extract date from date field for date-aware grouping
                date_field = analysis['date']
                if isinstance(date_field, str):
                    analysis_date = date_field.split(' ')[0]  # Get YYYY-MM-DD part
                else:
                    analysis_date = date_field.strftime('%Y-%m-%d')
                
                location_key = f"{analysis['salle']}-{analysis['matiere']}-{analysis_date}"
                batch = analysis.get('batch', 1)
                if batch is None:
                    batch = 1
                
                if batch == location_batches[location_key]:
                    filtered_general_analyses.append(analysis)
        
        # Apply correct batch logic for mobility analyses
        filtered_mobility_analyses = []
        
        if mobility_analyses_raw:
            # Group analyses by location+CNE+day
            location_groups = {}
            
            for analysis in mobility_analyses_raw:
                # Extract date from timestamp for date-aware grouping
                timestamp = analysis['timestamp']
                if isinstance(timestamp, str):
                    analysis_date = timestamp.split(' ')[0]  # Get YYYY-MM-DD part
                else:
                    analysis_date = timestamp.strftime('%Y-%m-%d')
                
                location_key = f"{analysis['salle']}-{analysis['matiere']}-{analysis['cne']}-{analysis_date}"
                
                if location_key not in location_groups:
                    location_groups[location_key] = {
                        'verified': [],
                        'unverified': []
                    }
                
                if analysis.get('verification_action') in ['confirm', 'deny']:
                    # Add to verified list - keep all verified batches
                    location_groups[location_key]['verified'].append(analysis)
                else:
                    # Add to unverified list - will filter to highest batch only
                    location_groups[location_key]['unverified'].append(analysis)
            
                         # For each location group, keep all verified + highest unverified batch only if it's newer
            for location_key, group in location_groups.items():
                 # Add all verified analyses (permanent)
                 filtered_mobility_analyses.extend(group['verified'])
                 
                 # Find the highest verified batch number
                 highest_verified_batch = 0
                 if group['verified']:
                     for verified_analysis in group['verified']:
                         verified_batch = verified_analysis.get('batch', 1)
                         if verified_batch is None:
                             verified_batch = 1
                         if verified_batch > highest_verified_batch:
                             highest_verified_batch = verified_batch
                 
                 # For unverified analyses, keep only the highest batch IF it's higher than highest verified
                 if group['unverified']:
                     highest_unverified_batch = 0
                     highest_unverified_analysis = None
                     
                     for analysis in group['unverified']:
                         batch = analysis.get('batch', 1)
                         if batch is None:
                             batch = 1
                         
                         if batch > highest_unverified_batch:
                             highest_unverified_batch = batch
                             highest_unverified_analysis = analysis
                     
                     # Only add unverified if it's newer than the newest verified batch
                     if highest_unverified_analysis and highest_unverified_batch > highest_verified_batch:
                         filtered_mobility_analyses.append(highest_unverified_analysis)
        
        return {
            "general_analyses": filtered_general_analyses,
            "mobility_analyses": filtered_mobility_analyses
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_chef_center_analyses: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la récupération des analyses"
        )
@app.get("/health")
def health_check():
    """Health check endpoint with database pool status"""
    try:
        # Test database connection
        test_result = db.execute_query("SELECT 1 as test")
        
        if test_result is None:
            return {
                "status": "unhealthy",
                "database": "disconnected",
                "error": "Database query failed",
                "timestamp": get_morocco_time_str()
            }
        
        # Get pool status
        pool = db.engine.pool
        pool_status = {
            "pool_size": pool.size(),
            "checked_in": pool.checkedin(),
            "checked_out": pool.checkedout(),
            "overflow": pool.overflow(),
            "invalid": pool.invalid()
        }
        
        return {
            "status": "healthy",
            "database": "connected",
            "pool_status": pool_status,
            "timestamp": get_morocco_time_str()
        }
    except Exception as e:
        import traceback
        logger.error(f"Health check failed: {e}")
        return {
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e),
            "traceback": traceback.format_exc(),
            "timestamp": get_morocco_time_str()
        }

@app.post("/chef/centre/{code_centre}/api/verify")
async def verify_mobility_analysis(code_centre: str, verification_data: dict):
    """Confirm or deny mobility analysis"""
    try:
        id_analyse_mobilite = verification_data.get("id_analyse_mobilite")
        action = verification_data.get("action")  # "confirm" or "deny"
        
        if not id_analyse_mobilite or not action:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="ID d'analyse et action requis"
            )
        
        if action not in ["confirm", "deny"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Action doit être 'confirm' ou 'deny'"
            )
        
        # Insert verification record with GMT+1 timestamp
        current_timestamp = get_morocco_time_str()
        insert_query = """
        INSERT INTO analyse_verifier (id_analyse_mobilite_final, timestamp, action)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE action = %s, timestamp = %s
        """
        
        result = db.execute_update(insert_query, (id_analyse_mobilite, current_timestamp, action, action, current_timestamp))
        
        if not result:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Erreur lors de la vérification"
            )
        
        # Get analysis details for broadcast
        analysis_query = """
        SELECT amf.*, v.aref, v.dp, v.ville, v.lycee, v.salle, v.matiere, v.full_name as verificateur_name
        FROM analyse_mobility_final amf
        JOIN verificateur v ON amf.id_verificateur = v.id_verificateur
        WHERE amf.id_analyse = %s
        """
        
        analysis_details = db.execute_query(analysis_query, (id_analyse_mobilite,))
        
        if analysis_details:
            analysis = analysis_details[0]
            # Broadcast verification via WebSocket
            verification_broadcast = {
                "event": "verification_update",
                "type": "analyse_mobilite_verified",
                "id_analyse": id_analyse_mobilite,
                "action": action,
                "timestamp": current_timestamp,
                "data": {
                    "id": analysis["id_analyse"],
                    "date": analysis["timestamp"],
                    "cne": analysis["cne"],
                    "aref": analysis["aref"],
                    "dp": analysis["dp"],
                    "ville": analysis["ville"],
                    "lycee": analysis["lycee"],
                    "salle": analysis["salle"],
                    "matiere": analysis["matiere"],
                    "verificateur_name": analysis["verificateur_name"],
                    "operateur": analysis.get("operator", "ORANGE"),  # Add operator field
                    "operator": analysis.get("operator", "ORANGE"),  # Alternative field name
                    "verification_action": action,
                    "batch": analysis.get("batch", 1)  # Add batch field
                }
            }
            
            # Broadcast to all connected clients
            await manager.broadcast(json.dumps(verification_broadcast))
        
        return {
            "status": "success",
            "message": f"Analyse {'confirmée' if action == 'confirm' else 'refusée'}",
            "action": action,
            "id_analyse": id_analyse_mobilite
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in verify_mobility_analysis: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la vérification"
        )

@app.get("/test/verified")
def test_verified():
    return {"message": "Test endpoint working"}

@app.get("/favicon.ico")
def get_favicon():
    return FileResponse("front-end/logo-cropped.svg")

@app.get("/static/logo-cropped.svg")
def get_logo_cropped():
    return FileResponse("front-end/logo-cropped.svg", media_type="image/svg+xml")

@app.get("/static/logo.svg")
def get_logo():
    return FileResponse("front-end/logo.svg", media_type="image/svg+xml")

@app.get("/logo-cropped.svg")
def get_logo_cropped_root():
    return FileResponse("front-end/logo-cropped.svg", media_type="image/svg+xml")

@app.get("/logo.svg")
def get_logo_root():
    return FileResponse("front-end/logo.svg", media_type="image/svg+xml")


@app.get("/t3shield/api/download-configuration")
def download_configuration():
    """
    Download configuration.json file with data fetched from database
    """
    try:
        # Fetch data from database tables
        db_instance = Database()
        db_instance.connect()
        
        # Get examen data
        examen_query = "SELECT examen FROM examen ORDER BY examen"
        examen_results = db_instance.execute_query(examen_query)
        examen_list = [row['examen'] for row in examen_results] if examen_results else []
        
        # Get salle data with numeric sorting
        salle_query = """
        SELECT salle FROM salle 
        ORDER BY CAST(SUBSTRING(salle, LENGTH('Salle ') + 1) AS UNSIGNED)
        """
        salle_results = db_instance.execute_query(salle_query)
        salle_list = [row['salle'] for row in salle_results] if salle_results else []
        
        # Get matiere (session) data with numeric sorting
        matiere_query = """
        SELECT session FROM matiere 
        ORDER BY CAST(SUBSTRING(session, LENGTH('Session ') + 1) AS UNSIGNED)
        """
        matiere_results = db_instance.execute_query(matiere_query)
        matiere_list = [row['session'] for row in matiere_results] if matiere_results else []
        
        db_instance.close()
        
        # Create the configuration data structure
        config_data = {
            "Examen": examen_list,
            "Salle": salle_list,
            "Matiere": matiere_list
        }
        
        # Convert to JSON string
        config_json = json.dumps(config_data, indent=2, ensure_ascii=False)
        
        # Create file-like object
        file_buffer = BytesIO(config_json.encode('utf-8'))
        file_buffer.seek(0)
        
        return StreamingResponse(
            iter([file_buffer.getvalue()]),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=configuration.json"}
        )
        
    except Exception as e:
        if 'db_instance' in locals():
            db_instance.close()
        raise HTTPException(status_code=500, detail=f"Error generating configuration file: {str(e)}")

@app.get("/t3shield/api/download-aref-centres")
def download_aref_centres():
    """
    Download aref_province_centres.xlsx file
    """
    try:
        # Get data from database for the Excel file
        db = Database()
        db.connect()
        
        query = """
        SELECT 
            a.codeAref,
            d.name as province_name,
            SUBSTRING_INDEX(l.name_fr, ' (', 1) as centre_label_fr,
            SUBSTRING_INDEX(l.name_ar, ' (', 1) as centre_label_ar,
            l.codeCentre
        FROM lycee l
        JOIN ville v ON l.id_ville = v.id_ville
        JOIN dp d ON v.id_dp = d.id_dp
        JOIN aref a ON d.codeAref = a.codeAref
        ORDER BY a.codeAref, d.name, l.name_fr
        """
        
        results = db.execute_query(query)
        
        # Convert to DataFrame
        df = pd.DataFrame(results)
        
        # Rename columns to match expected format
        df = df.rename(columns={
            'codeAref': 'codeAref',
            'province_name': 'Province',
            'centre_label_fr': 'Centre Label Fr',
            'centre_label_ar': 'Centre Label Ar',
            'codeCentre': 'codeCentre'
        })
        
        # Create Excel file in memory
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Centres', index=False)
        
        excel_buffer.seek(0)
        
        return StreamingResponse(
            iter([excel_buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=aref_province_centres.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel file: {str(e)}")
    
    finally:
        db.close()

# Configuration management endpoints
@app.get("/t3shield/api/configuration/{config_type}")
def get_configuration_items(config_type: str, request: Request = None):
    """Get configuration items for a specific type (examen, salle, matiere)"""
    try:
        # Validate config_type
        valid_types = {
            'examen': ('examen', 'examen'),
            'salle': ('salle', 'salle'), 
            'matiere': ('matiere', 'session')
        }
        
        if config_type not in valid_types:
            raise HTTPException(status_code=400, detail="Invalid configuration type")
        
        table_name, column_name = valid_types[config_type]
        
        db_instance = Database()
        db_instance.connect()
        
        # Build query with proper sorting based on config type
        if config_type == 'salle':
            query = f"""
            SELECT {column_name} FROM {table_name} 
            ORDER BY CAST(SUBSTRING({column_name}, LENGTH('Salle ') + 1) AS UNSIGNED)
            """
        elif config_type == 'matiere':
            query = f"""
            SELECT {column_name} FROM {table_name} 
            ORDER BY CAST(SUBSTRING({column_name}, LENGTH('Session ') + 1) AS UNSIGNED)
            """
        else:
            # For examen, use regular alphabetical sorting
            query = f"SELECT {column_name} FROM {table_name} ORDER BY {column_name}"
        
        results = db_instance.execute_query(query)
        
        db_instance.close()
        
        items = [row[column_name] for row in results] if results else []
        
        return {
            "type": config_type,
            "items": items,
            "count": len(items)
        }
        
    except Exception as e:
        if 'db_instance' in locals():
            db_instance.close()
        raise HTTPException(status_code=500, detail=f"Error retrieving {config_type} items: {str(e)}")

@app.post("/t3shield/api/configuration/{config_type}")
def add_configuration_item(config_type: str, item_data: dict, request: Request = None):
    """Add a new configuration item"""
    try:
        # Check authentication and permissions
        user = get_current_user(request)
        if not user or user["profile_id"] != 1:  # Only super admin can modify
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Validate config_type and get item name
        valid_types = {
            'examen': ('examen', 'examen'),
            'salle': ('salle', 'salle'), 
            'matiere': ('matiere', 'session')
        }
        
        if config_type not in valid_types:
            raise HTTPException(status_code=400, detail="Invalid configuration type")
        
        item_name = item_data.get('name', '').strip()
        if not item_name:
            raise HTTPException(status_code=400, detail="Item name is required")
        
        table_name, column_name = valid_types[config_type]
        
        db_instance = Database()
        db_instance.connect()
        
        # Check if item already exists
        check_query = f"SELECT {column_name} FROM {table_name} WHERE {column_name} = %s"
        existing = db_instance.execute_query(check_query, (item_name,))
        
        if existing:
            db_instance.close()
            raise HTTPException(status_code=400, detail=f"{config_type} '{item_name}' already exists")
        
        # Insert new item
        insert_query = f"INSERT INTO {table_name} ({column_name}) VALUES (%s)"
        result = db_instance.execute_update(insert_query, (item_name,))
        
        db_instance.close()
        
        if result:
            return {"message": f"{config_type} '{item_name}' added successfully"}
        else:
            raise HTTPException(status_code=500, detail=f"Failed to add {config_type}")
        
    except HTTPException:
        raise
    except Exception as e:
        if 'db_instance' in locals():
            db_instance.close()
        raise HTTPException(status_code=500, detail=f"Error adding {config_type}: {str(e)}")

@app.delete("/t3shield/api/configuration/{config_type}/{item_name}")
def delete_configuration_item(config_type: str, item_name: str, request: Request = None):
    """Delete a configuration item"""
    try:
        # Check authentication and permissions
        user = get_current_user(request)
        if not user or user["profile_id"] != 1:  # Only super admin can modify
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Validate config_type
        valid_types = {
            'examen': ('examen', 'examen'),
            'salle': ('salle', 'salle'), 
            'matiere': ('matiere', 'session')
        }
        
        if config_type not in valid_types:
            raise HTTPException(status_code=400, detail="Invalid configuration type")
        
        table_name, column_name = valid_types[config_type]
        
        db_instance = Database()
        db_instance.connect()
        
        # Check if item exists
        check_query = f"SELECT {column_name} FROM {table_name} WHERE {column_name} = %s"
        existing = db_instance.execute_query(check_query, (item_name,))
        
        if not existing:
            db_instance.close()
            raise HTTPException(status_code=404, detail=f"{config_type} '{item_name}' not found")
        
        # Delete item
        delete_query = f"DELETE FROM {table_name} WHERE {column_name} = %s"
        result = db_instance.execute_update(delete_query, (item_name,))
        
        db_instance.close()
        
        if result:
            return {"message": f"{config_type} '{item_name}' deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail=f"Failed to delete {config_type}")
        
    except HTTPException:
        raise
    except Exception as e:
        if 'db_instance' in locals():
            db_instance.close()
        raise HTTPException(status_code=500, detail=f"Error deleting {config_type}: {str(e)}")

@app.get("/t3shield/api/cheat-rates")
def get_cheat_rates(
    start_date: Optional[str] = None, 
    end_date: Optional[str] = None,
    codeCentre: Optional[str] = None,
    aref: Optional[str] = None,
    request: Request = None
    ):
    """
    Get cheat rate data with optional filtering
    """
    try:
        # Get user AREF filter
        user_aref = get_user_aref_filter(request) if request else None
        
        # If user is AREF admin, use their AREF code instead of the provided one
        if user_aref:
            aref = user_aref
        
        # Build the base query
        query_parts = [
            "SELECT cr.*, l.name_fr as lycee_name, l.name_ar as lycee_name_ar",
            "FROM cheat_rate cr",
            "JOIN lycee l ON cr.codeCentre = l.codeCentre",
            "WHERE 1=1"
        ]
        params = []
        
        # Add date filters if provided
        if start_date:
            query_parts.append("AND cr.date >= %s")
            params.append(start_date)
        
        if end_date:
            query_parts.append("AND cr.date <= %s")
            params.append(end_date)
        
        # Add center filter if provided
        if codeCentre:
            query_parts.append("AND cr.codeCentre = %s")
            params.append(codeCentre)
        
        # Add aref filter if provided
        if aref:
            query_parts.append("AND l.id_ville IN (SELECT id_ville FROM ville WHERE id_dp IN (SELECT id_dp FROM dp WHERE codeAref = %s))")
            params.append(aref)
        
        # Order by date descending
        query_parts.append("ORDER BY cr.date DESC")
        
        # Execute the query
        query = " ".join(query_parts)
        results = db.execute_query(query, tuple(params))
        
        if results is None:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to retrieve cheat rate data"
            )
        
        # Calculate additional statistics
        total_records = len(results)
        total_students = sum(record['nbr_etudiant'] for record in results)
        total_detections = sum(record['nbr_detection'] for record in results)
        overall_rate = (total_detections / total_students * 100) if total_students > 0 else 0
        
        # Transform the data
        cheat_rates = []
        for record in results:
            rate = (record['nbr_detection'] / record['nbr_etudiant'] * 100) if record['nbr_etudiant'] > 0 else 0
            cheat_rates.append({
                "taux_id": record['taux_id'],
                "codeCentre": record['codeCentre'],
                "lycee_name": record['lycee_name'],
                "lycee_name_ar": record['lycee_name_ar'],
                "salle": record['salle'],
                "matiere": record['matiere'],
                "examen": record['examen'],
                "date": record['date'],
                "nbr_etudiant": record['nbr_etudiant'],
                "nbr_detection": record['nbr_detection'],
                "taux_cheat": round(rate, 2)
            })
        
        return {
            "data": cheat_rates,
            "summary": {
                "total_records": total_records,
                "total_students": total_students,
                "total_detections": total_detections,
                "overall_cheat_rate": round(overall_rate, 2)
            }
        }
        
    except Exception as e:
        logger.error(f"Error retrieving cheat rates: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving cheat rate data: {str(e)}"
        )

@app.get("/t3shield/api/cheat-rates-summary")
def get_cheat_rates_summary(
    year: Optional[int] = None,
    aref: Optional[str] = None,
    request: Request = None
    ):
    try:
        db = Database()
        
        # Get user AREF filter
        user_aref = get_user_aref_filter(request) if request else None
        
        # If user is AREF admin, use their AREF code instead of the provided one
        if user_aref:
            aref = user_aref
        
        # Build the base query
        query = """
            SELECT 
                YEAR(date_creation) as year,
                COUNT(*) as total_analyses,
                SUM(CASE WHEN verification_status = 'confirmed' THEN 1 ELSE 0 END) as confirmed_cheating,
                ROUND(
                    (SUM(CASE WHEN verification_status = 'confirmed' THEN 1 ELSE 0 END) / COUNT(*)) * 100, 
                    2
                ) as cheat_rate_percentage
            FROM cheat_rate 
            WHERE 1=1
        """
        params = []
        
        if year:
            query += " AND YEAR(date_creation) = %s"
            params.append(year)
            
        if aref:
            query += " AND aref = %s"
            params.append(aref)
            
        query += " GROUP BY YEAR(date_creation) ORDER BY year DESC"
        
        results = db.execute_query(query, params)
        
        if results is None:
            return {"error": "Database query failed"}
            
        return {
            "success": True,
            "data": results
        }
        
    except Exception as e:
        return {"error": f"Failed to get cheat rates summary: {str(e)}"}

@app.get("/t3shield/api/geojson/regions")
def get_regions_geojson():
    """Fetch regions GeoJSON data from the database."""
    try:
        db = Database()
        
        # Query to get all regions with their polygon data
        query = """
            SELECT 
                id_aref,
                codeAref,
                name,
                abbreviation,
                coordinates,
                region_polygone
            FROM aref 
            WHERE region_polygone IS NOT NULL 
            AND region_polygone != ''
            ORDER BY codeAref
            """
        
        results = db.execute_query(query)
        
        if results is None:
            return {"error": "Database query failed"}
        
        # Create a FeatureCollection from the database results
        features = []
        for row in results:
            try:
                # Parse the region polygon data
                if row['region_polygone']:
                    region_data = json.loads(row['region_polygone'])
                    
                    # Check if region_data is a Feature object and extract geometry
                    if isinstance(region_data, dict):
                        if region_data.get('type') == 'Feature':
                            # Extract geometry from Feature object
                            geometry = region_data.get('geometry')
                        elif region_data.get('type') == 'FeatureCollection':
                            # Extract first feature's geometry from FeatureCollection
                            features_list = region_data.get('features', [])
                            if features_list:
                                geometry = features_list[0].get('geometry')
                            else:
                                continue
                        else:
                            # Assume it's already a geometry object
                            geometry = region_data
                        
                        # Create a feature for this region
                        feature = {
                            "type": "Feature",
                            "geometry": geometry,
                            "properties": {
                                "ID_Reg": row['codeAref'],
                                "Nom_Region": row['name'],
                                "Abbreviation": row['abbreviation'],
                                "id_aref": row['id_aref']
                            }
                        }
                        features.append(feature)
            except json.JSONDecodeError as e:

                continue
            except Exception as e:

                continue
        
        geojson_data = {
            "type": "FeatureCollection",
            "features": features
        }
        
        return geojson_data
        
    except Exception as e:
        return {"error": f"Failed to get regions GeoJSON: {str(e)}"}

@app.get("/t3shield/api/geojson/provinces")
def get_provinces_geojson():
    """Fetch provinces GeoJSON data from the database."""
    try:
        db = Database()
        
        # Query to get all regions with their provinces polygon data
        query = """
            SELECT 
                id_aref,
                codeAref,
                name,
                abbreviation,
                provinces_polygone
            FROM aref 
            WHERE provinces_polygone IS NOT NULL 
            AND provinces_polygone != ''
            ORDER BY codeAref
        """
        
        results = db.execute_query(query)
        
        if results is None:
            return {"error": "Database query failed"}
        
        # Create a FeatureCollection from the database results
        features = []
        for row in results:
            try:
                # Parse the provinces polygon data
                if row['provinces_polygone']:
                    provinces_data = json.loads(row['provinces_polygone'])
                    
                    # Add all features from this region's provinces
                    if 'features' in provinces_data:
                        for feature in provinces_data['features']:
                            # Ensure the feature has the correct region code
                            if 'properties' in feature:
                                feature['properties']['CodeRegi_1'] = row['codeAref']
                                feature['properties']['Region_Name'] = row['name']
                            features.append(feature)
            except json.JSONDecodeError as e:

                continue
        
        geojson_data = {
            "type": "FeatureCollection",
            "features": features
        }
        
        return geojson_data
        
    except Exception as e:
        return {"error": f"Failed to get provinces GeoJSON: {str(e)}"}

@app.get("/t3shield/api/geojson/provinces/{region_id}")
def get_provinces_for_region(region_id: str):
    """Fetch provinces GeoJSON data for a specific region from the database."""
    try:
        db = Database()
        
        # Query to get provinces for a specific region
        query = """
            SELECT 
                id_aref,
                codeAref,
                name,
                abbreviation,
                provinces_polygone
            FROM aref 
            WHERE codeAref = %s 
            AND provinces_polygone IS NOT NULL 
            AND provinces_polygone != ''
        """
        
        results = db.execute_query(query, (region_id,))
        
        if results is None or not results:
            return {"error": "Region not found or no provinces data available"}
        
        row = results[0]
        
        try:
            # Parse the provinces polygon data
            if row['provinces_polygone']:
                provinces_data = json.loads(row['provinces_polygone'])
                
                # Ensure all features have the correct region code
                if 'features' in provinces_data:
                    for feature in provinces_data['features']:
                        if 'properties' in feature:
                            feature['properties']['CodeRegi_1'] = row['codeAref']
                            feature['properties']['Region_Name'] = row['name']
                
                return provinces_data
            else:
                return {"error": "No provinces data available for this region"}
                
        except json.JSONDecodeError as e:
            return {"error": f"Error parsing provinces data: {str(e)}"}
        
    except Exception as e:
        return {"error": f"Failed to get provinces for region: {str(e)}"}